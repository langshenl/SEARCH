#!/usr/bin/env python3
import json
import re
import sys
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    print('missing dependency: openpyxl', file=sys.stderr)
    sys.exit(2)

COLUMNS = ["标题", "正文", "摘要", "发文机关", "发布时间", "原始链接", "关键词", "类型", "地区"]
TYPE_RULES = [
    ('项目申报', ['申报', '申报指南', '申报条件']),
    ('发展规划', ['发展规划', '规划纲要', '规划']),
    ('工作部署', ['工作要点', '工作方案', '通知', '召开', '部署']),
    ('产业扶持', ['扶持', '支持', '奖补', '补贴', '若干措施', '支持措施']),
    ('管理规范', ['管理办法', '办法', '细则', '监管']),
    ('建设实施', ['建设方案', '建设', '实施']),
    ('政策解读', ['解读', '答记者问']),
    ('统计报告', ['统计公报', '工作报告', '报告', '执行情况']),
    ('资金预算', ['预算', '资金', '财政']),
    ('人才引进', ['人才', '引进']),
]
ORG_PATTERNS = [
    r"([^\n，。；]{2,40}(?:人民政府|人民法院|人民检察院|委员会|管理局|监管局|发展和改革委员会|发展改革委|财政厅|财政局|教育厅|教育局|商务厅|商务局|文旅厅|文化和旅游厅|博物馆|办公室))",
]
DATE_PATTERNS = [
    r"(20\d{2}[-/.年]\d{1,2}[-/.月]\d{1,2}日?)",
    r"(20\d{2}年\d{1,2}月\d{1,2}日)",
    r"(20\d{2}-\d{2}-\d{2})",
]


def normalize_date(date_text, expected_year=''):
    if not date_text:
        return ''
    date_text = str(date_text).strip()
    m = re.search(r'(20\d{2})', date_text)
    if expected_year and (not m or m.group(1) != expected_year):
        return ''
    m2 = re.search(r'(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})', date_text)
    if m2:
        y, mo, d = m2.groups()
        return f'{y}-{int(mo):02d}-{int(d):02d}'
    m3 = re.search(r'(20\d{2})-(\d{2})-(\d{2})', date_text)
    if m3:
        return f'{m3.group(1)}-{m3.group(2)}-{m3.group(3)}'
    return date_text if (not expected_year or expected_year in date_text) else ''


def first_sentence(text, max_len=100):
    if not text:
        return ''
    for part in re.split(r'[。！？\n]+', text):
        part = part.strip()
        if len(part) >= 8 and not any(w in part.lower() for w in [
            '当前位置', 'header', 'menu', '导航', '网站地图', '版权', 'copyright'
        ]):
            return part[:max_len]
    return ''


def first_two_sentences(text, max_len=160):
    if not text:
        return ''
    parts = []
    for part in re.split(r'[。！？\n]+', text):
        part = part.strip()
        if len(part) >= 8 and not any(w in part.lower() for w in [
            '当前位置', 'header', 'menu', '导航', '网站地图', '版权', 'copyright'
        ]):
            parts.append(part)
        if len(parts) >= 2:
            break
    return '。'.join(parts)[:max_len]


def is_url_valid(url, timeout=5):
    """方案一：先按 URL 规则硬过滤，再做弱校验"""
    if not url or not url.startswith(('http://', 'https://')):
        return False

    low = url.lower()
    # 列表页 / 索引页
    if any(x in low for x in ['/list', '/index', '_index', '-index']):
        return False
    # 附件链接
    if any(low.endswith('.' + x) or f'.{x}?' in low or f'.{x}' in low for x in ['pdf', 'doc', 'docx', 'ofd', 'xls', 'xlsx', 'zip', 'rar', '7z', 'tar', 'gz']):
        return False

    try:
        req = urllib.request.Request(url)
        req.add_header('User-Agent', 'Mozilla/5.0')
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            code = resp.getcode()
            if code in (404, 410):
                return False
            content = resp.read(100000).decode('utf-8', errors='ignore')
            low = content.lower()
            title_match = re.search(r'<title>(.*?)</title>', content, re.I | re.S)
            page_title = title_match.group(1).strip().lower() if title_match else ''
            error_signals = ['404 not found', '页面不存在', '您访问的页面不存在', 'not found', '内容不存在', '信息不存在', '已删除', '已失效']
            if any(k in low for k in error_signals) or any(k in page_title for k in error_signals):
                return False
            return True
    except urllib.error.HTTPError as e:
        if e.code in (404, 410):
            return False
        return True
    except Exception:
        return True


def first_match(patterns, text):
    for p in patterns:
        m = re.search(p, text or '')
        if m:
            return m.group(1).strip()
    return ''


def infer_type(title, summary, body):
    text = f'{title or ""} {summary or ""} {body or ""}'
    for label, keywords in TYPE_RULES:
        if any(k in text for k in keywords):
            return label
    return '综合管理'


def infer_region(query, provided_region, text):
    if provided_region:
        return provided_region
    provinces = ["湖北", "湖南", "广东", "浙江", "江苏", "北京", "上海", "天津", "重庆", "四川", "河南", "河北", "山东", "山西", "陕西", "安徽", "福建", "江西", "广西", "云南", "贵州", "海南", "辽宁", "吉林", "黑龙江", "内蒙古", "宁夏", "新疆", "西藏", "青海", "甘肃"]
    merged = ' '.join([query or '', text or ''])
    for p in provinces:
        if p in merged:
            return p
    return '全国'



def clean_text(text, summary=''):
    """加强正文清洗：去 script/style/html 噪音，保留可读文本"""
    if not text:
        return ''

    text = re.sub(r'!\[[^\]]*\]\([^\)]*\)', ' ', text)
    text = re.sub(r'#:~:text=[^\s\)]*', '', text)
    text = re.sub(r'<script[\s\S]*?</script>', ' ', text, flags=re.I)
    text = re.sub(r'<style[\s\S]*?</style>', ' ', text, flags=re.I)
    text = re.sub(r'<noscript[\s\S]*?</noscript>', ' ', text, flags=re.I)
    text = re.sub(r'<!--[\s\S]*?-->', ' ', text)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'&[a-zA-Z]{2,10};', ' ', text)
    text = re.sub(r'&#\d+;', ' ', text)
    text = re.sub(r'&#x[0-9a-fA-F]+;', ' ', text)

    cut_markers = ['您访问的链接将离开', '是否继续', '离开“', '离开本站', '门户网站']
    for marker in cut_markers:
        idx = text.find(marker)
        if idx != -1:
            text = text[:idx]
            break

    text = re.sub(r'\s+', ' ', text)

    noise_patterns = [
        r'\bvar\s+\$?\w+\s*=.*?;',
        r'\$\.ajax\(|jQuery\(|\$\(function',
        r'https?://[^\s]{5,200}',
        r'\bfunction\s*\w+\s*\([^)]*\)\s*\{',
    ]
    for p in noise_patterns:
        text = re.sub(p, ' ', text, flags=re.I)

    parts = re.split(r'(?<=[。！？；])\s*', text)
    cleaned = []
    seen = set()
    for part in parts:
        part = part.strip()
        if len(part) < 8:
            continue
        if any(w in part.lower() for w in [
            'copyright', '版权所有', 'cookie', '隐私政策', '网站地图',
            '请您登录', '请登录', '无障碍', '长者专区', 'loading',
            'ajax', 'javascript', 'jquery', 'header 开始', 'header 结束',
            'content 开始', '当前位置 开始', '当前位置 结束', '移动端菜单开始', '移动端菜单结束'
        ]):
            continue
        key = part[:40].strip().lower()
        if key and key not in seen:
            seen.add(key)
            cleaned.append(part)
    return '\n'.join(cleaned)


def normalize_item(item, query, region):
    expected_year_match = re.search(r'(20\d{2})', query or '')
    expected_year = expected_year_match.group(1) if expected_year_match else ''

    url = (item.get('url') or item.get('id') or '').strip()
    url = re.sub(r'#.*$', '', url)
    exa_summary = (item.get('summary') or item.get('snippet') or '').strip()
    raw_text = (item.get('text') or item.get('content') or '').strip()
    raw_title = (item.get('title') or '').strip()

    # 方案3：此处不负责过滤（过滤在 main() 做），只做清洗
    effective_text = raw_text
    text = clean_text(effective_text, exa_summary)

    summary = first_sentence(text, max_len=120)
    if len(summary) < 50:
        summary = first_two_sentences(text, max_len=160) or summary
    if not summary:
        summary = exa_summary or first_two_sentences(raw_text, max_len=160) or first_sentence(raw_text, max_len=120)

    title = (item.get('title') or '').strip()
    if not title:
        title = first_sentence(text) or first_sentence(summary) or '标题缺失'

    # 标题去噪声：去掉页面顶部导航残留词
    title = re.sub(r'^(政府工作报告|页面不存在|404).*?[-–—]\s*', '', title)

    org = first_match(ORG_PATTERNS, '\n'.join([title, summary, text]))

    exa_date = (
        item.get('publishedDate')
        or item.get('published_date')
        or item.get('publishDate')
        or item.get('publish_date')
        or ''
    )
    publish_date = normalize_date(exa_date, expected_year)
    if not publish_date:
        body_date = first_match(DATE_PATTERNS, '\n'.join([title, summary, text]))
        publish_date = normalize_date(body_date, expected_year)

    row_region = infer_region(query, region, '\n'.join([title, summary, text]))
    doc_type = infer_type(title, summary, text)
    return {
        '标题': title,
        '正文': text,
        '摘要': summary,
        '发文机关': org,
        '发布时间': publish_date,
        '原始链接': url,
        '关键词': query,
        '类型': doc_type,
        '地区': row_region,
    }


def _summarize(rows, query):
    """生成学术风格的搜索结果摘要"""
    if not rows:
        return f"未找到与「{query}」相关的结果。"

    total = len(rows)

    # 按类型统计
    type_count = {}
    for r in rows:
        t = r.get('类型', '未知')
        type_count[t] = type_count.get(t, 0) + 1
    top_types = sorted(type_count.items(), key=lambda x: -x[1])[:6]

    # 合并所有正文和标题做主题分析
    all_text = ' '.join([r.get('标题', '') + ' ' + r.get('正文', '') for r in rows])

    # 推断政策主题（按关键词命中）
    themes = []
    theme_hits = {
        '新能源开发与利用': ['新能源', '光伏', '风电', '储能', '充电桩', '氢能', '生物质能'],
        '财政扶持与价格机制': ['电价', '补贴', '财政', '奖补', '预算', '资金', '税收', '优惠'],
        '项目建设与产业布局': ['项目', '建设', '投资', '产业', '园区', '基地', '集群'],
        '新能源汽车推广': ['新能源汽车', '电动汽车', '充电设施', '汽车下乡', '绿车'],
        '节能降碳与绿色发展': ['节能', '降碳', '碳达峰', '碳中和', '绿色', '排放', '减排'],
        '发展规划与实施方案': ['规划', '方案', '意见', '计划', '纲要', '要点'],
        '资源交易与市场机制': ['交易', '招标', '投标', '采购', '竞价', '市场'],
    }
    for theme, kws in theme_hits.items():
        if sum(1 for kw in kws if kw in all_text) >= 2:
            themes.append(theme)
    theme_str = '、'.join(themes[:4]) if themes else '综合性政策'

    lines = [
        f"本次检索共获得 {total} 项政策文件，涵盖以下主题：",
        f"",
        f"【内容主题】{theme_str}",
        f"",
        f"【文件类型】",
    ]
    for t, c in top_types:
        pct = c / total * 100
        lines.append(f"  · {t}：{c}项（{pct:.0f}%）")

    lines.append(f"")
    lines.append(f"【研究说明】上述文件主要围绕{themes[0] if themes else '政策体系建设'}等议题展开，从顶层规划、财政支持、项目实施等多维度呈现地方政策动态。")

    return '\n'.join(lines)


def main():
    if len(sys.argv) < 4:
        print('usage: write_exa_results.py <input_json> <query> <region>', file=sys.stderr)
        sys.exit(1)
    input_json = Path(sys.argv[1]).expanduser()
    query = sys.argv[2]
    region = sys.argv[3]
    data = json.loads(input_json.read_text(encoding='utf-8'))
    if isinstance(data, dict):
        items = data.get('results') or data.get('items') or []
    elif isinstance(data, list):
        items = data
    else:
        items = []

    valid_items = []
    invalid_count = 0
    empty_body_count = 0
    for item in items:
        url = (item.get('url') or item.get('id') or '').strip()
        if url and is_url_valid(url):
            valid_items.append(item)
        else:
            invalid_count += 1

    rows = [normalize_item(item, query, region) for item in valid_items]
    # 正文为空 或 正文字数<50 → 过滤
    rows_before = rows
    rows = [r for r in rows if len(r['正文'].strip()) >= 50]
    empty_body_count = len(rows_before) - len(rows)
    safe = re.sub(r'[^\w\u4e00-\u9fff-]+', '_', query)[:60].strip('_') or 'exa-search'
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_dir = Path('~/Desktop/exa搜索文件夹').expanduser() / f'{safe}_{ts}'
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = '搜索结果'
    ws.append(COLUMNS)

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    col_widths = {'标题': 30, '正文': 50, '摘要': 30, '发文机关': 20, '发布时间': 15, '原始链接': 60, '关键词': 20, '类型': 12, '地区': 12}
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if col_name == '原始链接':
                url = row[col_name]
                if url:
                    cell.hyperlink = url
                    cell.value = url
                    cell.font = Font(color='0563C1', underline='single')
            else:
                cell.value = row[col_name]
        ws.row_dimensions[row_idx].height = 60

    for col_idx, col_name in enumerate(COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths.get(col_name, 15)

    ws.row_dimensions[1].height = 22
    xlsx_path = out_dir / '搜索结果.xlsx'
    wb.save(xlsx_path)

    # 生成并打印自然语言总结
    summary_msg = _summarize(rows, query)
    print(summary_msg)
    print()

    print(json.dumps({
        'output_dir': str(out_dir),
        'excel': str(xlsx_path),
        'count': len(rows),
        'invalid_count': invalid_count,
        'empty_body_count': empty_body_count,
        'columns': COLUMNS,
    }, ensure_ascii=False))


if __name__ == '__main__':
    main()
