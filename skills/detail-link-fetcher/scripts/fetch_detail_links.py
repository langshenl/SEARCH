#!/usr/bin/env python3
from __future__ import annotations

import re
import sys
import time
import urllib.request
from html import unescape
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook

USER_AGENT = (
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
    'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36'
)
TIMEOUT = 20

FIELDS = [
    '政策ID', '标题', '正文', '摘要', '发文机关', '联合发文机关', '发布时间', '文号',
    '政策层级', '地区', '所属行业', '类型', '政策主题', '关键词', '适用对象', '支持方式',
    '申报条件', '申报时间', '有效期', '政策状态', '原始链接', '附件链接', '相似政策', '上位政策/下位政策关系',
    '抓取状态', '正文长度', '是否检测到附件', '是否需要浏览器兜底'
]


def fetch_text(url: str) -> str:
    req = urllib.request.Request(url, headers={'User-Agent': USER_AGENT, 'Accept-Language': 'zh-CN,zh;q=0.9'})
    with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
        charset = resp.headers.get_content_charset() or 'utf-8'
        return resp.read().decode(charset, errors='ignore')


def strip_html(text: str) -> str:
    text = re.sub(r'<script[\s\S]*?</script>', ' ', text, flags=re.I)
    text = re.sub(r'<style[\s\S]*?</style>', ' ', text, flags=re.I)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = unescape(text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def extract_meta(html: str, name: str) -> str:
    m = re.search(rf'<meta[^>]+(?:name|property)="{re.escape(name)}"[^>]+content="([^"]*)"', html, flags=re.I)
    return unescape(m.group(1)).strip() if m else ''


def extract_title(html: str) -> str:
    article_title = extract_meta(html, 'ArticleTitle')
    if article_title:
        return article_title
    m = re.search(r'<title>([\s\S]*?)</title>', html, flags=re.I)
    return strip_html(m.group(1)) if m else ''


def extract_body(html: str) -> str:
    patterns = [
        r'<div[^>]+class="[^"]*(?:article-content|article_content|articleCon|wp_articlecontent|TRS_Editor|content|article|zoom)[^"]*"[^>]*>([\s\S]*?)</div>',
        r'<div[^>]+id="[^"]*(?:article|content|zoom|main-content)[^"]*"[^>]*>([\s\S]*?)</div>',
        r'<article[^>]*>([\s\S]*?)</article>',
        r'<div[^>]+class="[^"]*(?:main|detail|pages_content|news_detail|txt_con)[^"]*"[^>]*>([\s\S]*?)</div>',
    ]
    best = ''
    for pat in patterns:
        for m in re.finditer(pat, html, flags=re.I):
            txt = strip_html(m.group(1))
            if len(txt) > len(best):
                best = txt
    return best


def extract_attachments(html: str) -> str:
    links = re.findall(r'<a[^>]+href="([^"]+)"[^>]*>(.*?)</a>', html, flags=re.I | re.S)
    hits = []
    for href, text in links:
        t = strip_html(text)
        if re.search(r'附件|下载|PDF|DOC|DOCX|OFD|XLS|XLSX', t, flags=re.I) or re.search(r'\.(pdf|doc|docx|ofd|xls|xlsx|zip)$', href, flags=re.I):
            hits.append(href)
    return '\n'.join(dict.fromkeys(hits))


def derive_policy_id(url: str) -> str:
    patterns = ['(t\\d+_\\d+)', '/([A-Za-z0-9_-]{16,})\\.shtml(?:\\?|$)', '/([a-f0-9]{24,})\\.shtml(?:\\?|$)', '/(\\d{10,})/', 'content[_-](\\d+)', '/info[_-](\\d+)', 'article[_-](\\d+)', 'show[_-](\\d+)', 'index[_-](\\d+)', 'page[_-](\\d+)', '[?&]id=(\\d+)', '[?&]id=(\\w+)', '[?&](?:newsId|articleId|contentId)=(\\w+)', '/(\\d{8,})/', '(\\d{8,}[-_]\\d+)', '/(\\d+)\\.html']
    for pat in patterns:
        m = re.search(pat, url, flags=re.I)
        if m:
            return m.group(1)
    return ''


def derive_region(url: str, html: str) -> str:
    region_map = {
        'hubei.gov.cn': '湖北省',
        'hunan.gov.cn': '湖南省',
        'hebei.gov.cn': '河北省',
        'yn.gov.cn': '云南省',
        'sc.gov.cn': '四川省',
        'yn.gov.cn': '云南省',
        'hunan.gov.cn': '湖南省',
    }
    for host, region in region_map.items():
        if host in url:
            return region
    site_name = extract_meta(html, 'SiteName')
    for region in ['湖北省', '湖南省', '河北省', '云南省', '四川省']:
        if region.replace('省', '') in site_name or region in site_name:
            return region
    return ''


def classify_type(title: str, summary: str, body: str) -> str:
    text = f'{title} {summary} {body}'
    rules = [
        ('项目申报', ['申报', '申报指南', '申报条件']),
        ('发展规划', ['发展规划', '规划纲要', '规划']),
        ('工作部署', ['工作要点', '工作方案', '通知', '召开', '部署']),
        ('产业扶持', ['扶持', '支持', '奖补', '补贴', '若干措施', '支持措施']),
        ('管理规范', ['管理办法', '办法', '细则', '监管']),
        ('建设实施', ['建设方案', '建设', '实施']),
        ('政策解读', ['解读', '答记者问']),
        ('统计报告', ['统计公报', '工作报告', '报告', '执行情况']),
        ('资金预算', ['预算', '资金', '财政']),
        ('人才引进', ['人才', '引进']),
    ]
    for label, keywords in rules:
        if any(k in text for k in keywords):
            return label
    return '综合管理'


def status_and_fallback(row: dict) -> tuple[str, str, str]:
    body_len = len((row.get('正文') or '').strip())
    has_attachment = '是' if (row.get('附件链接') or '').strip() else '否'
    poor = False
    if not row.get('标题'):
        poor = True
    if body_len < 120:
        poor = True
    if not row.get('发布时间'):
        poor = True
    status = 'GOOD'
    if poor and body_len > 0:
        status = 'PARTIAL'
    if poor and body_len == 0:
        status = 'POOR'
    need_browser = '是' if status in {'PARTIAL', 'POOR'} else '否'
    return status, str(body_len), need_browser if need_browser else '否'


def parse_detail(url: str) -> dict:
    row = {k: '' for k in FIELDS}
    row['原始链接'] = url
    try:
        html = fetch_text(url)
    except Exception:
        row['抓取状态'] = 'FETCH_ERROR'
        row['正文长度'] = '0'
        row['是否检测到附件'] = '否'
        row['是否需要浏览器兜底'] = '是'
        return row

    row['政策ID'] = derive_policy_id(url)
    row['标题'] = extract_title(html)
    row['摘要'] = extract_meta(html, 'Description')
    row['发文机关'] = extract_meta(html, 'ContentSource')
    row['发布时间'] = extract_meta(html, 'PubDate')
    row['关键词'] = extract_meta(html, 'Keywords')
    row['正文'] = extract_body(html)
    row['附件链接'] = extract_attachments(html)
    row['地区'] = derive_region(url, html)
    row['类型'] = classify_type(row['标题'], row['摘要'], row['正文'])
    row['是否检测到附件'] = '是' if row['附件链接'] else '否'

    status, body_len, need_browser = status_and_fallback(row)
    row['抓取状态'] = status
    row['正文长度'] = body_len
    row['是否需要浏览器兜底'] = need_browser
    return row


def read_links(md_path: Path):
    text = md_path.read_text(encoding='utf-8', errors='ignore')
    urls = re.findall(r'\((http[^)]+)\)', text)
    seen = set()
    ordered = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            ordered.append(u)
    return ordered


def write_excel(rows, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = '政策详情'
    ws.append(FIELDS)
    for row in rows:
        ws.append([row.get(f, '') for f in FIELDS])

    header_index = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
    for field in ['原始链接', '附件链接']:
        if field not in header_index:
            continue
        col = header_index[field]
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            value = str(cell.value or '').strip()
            if not value:
                continue
            if field == '附件链接' and '\n' in value:
                first = value.splitlines()[0].strip()
                if first.startswith('http'):
                    cell.value = first
                    cell.hyperlink = first
                    cell.style = 'Hyperlink'
                continue
            if value.startswith('http'):
                cell.hyperlink = value
                cell.style = 'Hyperlink'
    wb.save(out_path)


def derive_name_from_links(link_md: Path) -> str:
    text = link_md.read_text(encoding='utf-8', errors='ignore')
    urls = re.findall(r'\((http[^)]+)\)', text)
    if not urls:
        return '搜索结果'
    host_map = {
        'www.hubei.gov.cn': '湖北省政策',
        'www.hunan.gov.cn': '湖南政策',
        'www.hebei.gov.cn': '河北政策',
        'www.yn.gov.cn': '云南省政策',
        'www.sc.gov.cn': '四川政策',
    }
    for url in urls:
        for host, name in host_map.items():
            if host in url:
                return name
    return '搜索结果'


def main():
    keyword = sys.argv[1] if len(sys.argv) > 1 else ''
    link_md = Path.home() / 'Desktop' / '处理文件夹' / '原文链接.md'
    out_dir = Path.home() / 'Desktop' / 'detail-h5'
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    keyword_name = keyword if keyword else derive_name_from_links(link_md)
    safe_keyword = re.sub(r'[\\/:*?"<>|]+', '_', keyword_name)
    out_excel = out_dir / f'{safe_keyword}_{stamp}.xlsx'

    text = link_md.read_text(encoding='utf-8', errors='ignore')
    raw_urls = re.findall(r'\((http[^)]+)\)', text)
    urls = read_links(link_md)
    rows = []
    for url in urls:
        rows.append(parse_detail(url))
        time.sleep(1)

    if out_excel.exists():
        out_excel.unlink()
    write_excel(rows, out_excel)

    print(f'INPUT_MD={link_md}')
    print(f'RAW_LINK_COUNT={len(raw_urls)}')
    print(f'UNIQUE_LINK_COUNT={len(urls)}')
    print(f'COUNT={len(rows)}')
    print(f'EXCEL={out_excel}')


if __name__ == '__main__':
    main()
