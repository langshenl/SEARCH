#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import sys
import time
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from html import unescape
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook

# 格式库目录（相对于脚本位置）
SCRIPT_DIR = Path(__file__).parent
PROVINCE_RULES_DIR = SCRIPT_DIR.parent / 'references' / 'province_rules'

# 全省份格式规则（延迟加载）
_PROVINCE_RULES_CACHE = {}


def load_province_rules():
    """加载所有省份格式规则"""
    if _PROVINCE_RULES_CACHE:
        return _PROVINCE_RULES_CACHE
    if not PROVINCE_RULES_DIR.exists():
        return {}
    for json_file in PROVINCE_RULES_DIR.glob('*.json'):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                rule = json.load(f)
                domain = rule.get('domain', '')
                if domain:
                    _PROVINCE_RULES_CACHE[domain] = rule
        except Exception:
            pass
    return _PROVINCE_RULES_CACHE


def match_province_rule(url: str):
    """根据URL匹配省份规则"""
    rules = load_province_rules()
    for domain, rule in rules.items():
        if domain in url:
            return rule
    return None


def extract_body_by_rule(html: str, rule: dict) -> str:
    """根据省份规则提取正文

    支持两种模式：
    1. 多模板模式（templates）：根据 HTML 内容自动选择模板
    2. 直接模式：content_start + content_end 直接定位
    """
    # 优先检测多模板模式（如四川）
    templates = rule.get('templates', {})
    if templates:
        # 四川：根据 HTML 中是否存在 UCAPCONTENT 选择模板
        if 'cms' in templates and '<UCAPCONTENT' in html:
            rule = {**rule, **templates['cms']}
        elif 'ywwd' in templates:
            rule = {**rule, **templates['ywwd']}
        else:
            # 无匹配模板，返回空
            return ''

    content_start_tag = rule.get('content_start', '')
    content_end = rule.get('content_end', '')

    # 定位正文结束位置
    end_idx = html.find(content_end)
    if end_idx == -1:
        return ''

    # 定位正文开始位置
    start_idx = 0
    if content_start_tag:
        content_start_pos = html.find(content_start_tag)
        if content_start_pos != -1:
            if content_start_tag.startswith('<div'):
                # 容器标签：从标签起始位置开始
                start_idx = content_start_pos
            else:
                # 自定义/结束标签：从标签之后开始
                div_close = html.find('</div>', content_start_pos)
                if div_close != -1:
                    start_idx = div_close + len('</div>')
                    while start_idx < len(html) and html[start_idx] in ' \n\t':
                        start_idx += 1

    if start_idx >= end_idx:
        return ''

    content = html[start_idx:end_idx]

    # 按规则清洗HTML
    clean_rules = rule.get('clean_rules', {})
    remove_tags = clean_rules.get('remove_tags', ['script', 'style', 'img'])
    block_tags = clean_rules.get('block_tags', ['p', 'br'])

    # 移除指定标签
    for tag in remove_tags:
        content = re.sub(rf'<{tag}[^>]*>[\s\S]*?</{tag}>', '', content, flags=re.I)

    # 块级标签转换行
    for tag in block_tags:
        content = re.sub(rf'</{tag}>', '\n', content, flags=re.I)
        content = re.sub(rf'<{tag}[^>]*>', '\n', content, flags=re.I)

    # 去掉所有剩余标签
    content = re.sub(r'<[^>]+>', '', content)
    content = unescape(content)

    # 清理空白
    content = re.sub(r'[ \t]+', ' ', content)
    content = re.sub(r'\n{3,}', '\n\n', content)
    return content.strip()
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

USER_AGENT = (
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
    'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36'
)
TIMEOUT = 5

# ⚠️ 字段配置（固定9个，不允许增删）
FIELDS = [
    '标题', '正文', '摘要', '发文机关', '发布时间', '原始链接',
    '关键词', '类型', '地区'
]


def fetch_text(url: str) -> str:
    req = urllib.request.Request(url, headers={'User-Agent': USER_AGENT, 'Accept-Language': 'zh-CN,zh;q=0.9'})
    with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
        charset = resp.headers.get_content_charset() or 'utf-8'
        return resp.read().decode(charset, errors='ignore')


def strip_html(text: str) -> str:
    text = re.sub(r'<script[\s\S]*?</script>', ' ', text, flags=re.I)
    text = re.sub(r'<style[\s\S]*?</style>', ' ', text, flags=re.I)
    # 保留块级标签为换行，不转为空格
    text = re.sub(r'<(?:br\s*/?|hr)[^>]*>', '\n', text, flags=re.I)
    text = re.sub(r'<(?:p|div|li|tr|section|article|header|footer|main|aside)[^>]*>', '\n', text, flags=re.I)
    # 再去掉剩余标签
    text = re.sub(r'<[^>]+>', ' ', text)
    text = unescape(text)
    # 清理换行：多个连续换行合并为一个空格/换行
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def extract_meta(html: str, name: str) -> str:
    m = re.search(rf'<meta[^>]+(?:name|property)="{re.escape(name)}"[^>]+content="([^"]*)"', html, flags=re.I)
    return unescape(m.group(1)).strip() if m else ''


def extract_title(html: str) -> str:
    article_title = extract_meta(html, 'ArticleTitle')
    if article_title:
        return article_title
    # Fallback：从 og:title 获取
    og_title = re.search(r'<meta[^>]+property="og:title"[^>]+content="([^"]*)"', html, flags=re.I)
    if og_title:
        return unescape(og_title.group(1)).strip()
    m = re.search(r'<title>([\s\S]*?)</title>', html, flags=re.I)
    return strip_html(m.group(1)) if m else ''


def extract_body(html: str, url: str = '') -> str:
    # 优先尝试格式库规则匹配
    if url:
        rule = match_province_rule(url)
        if rule:
            body_by_rule = extract_body_by_rule(html, rule)
            if len(body_by_rule) >= 100:
                return body_by_rule

    # 主 pattern（按优先级排列，越靠前优先级越高）
    main_patterns = [
        # 标准正文容器
        r'<div[^>]+class="[^"]*(?:article-content|article_content|articleCon|TRS_Editor)[^"]*"[^>]*>([\s\S]*?)</div>',
        r'<article[^>]*>([\s\S]*?)</article>',
        # id 命中
        r'<div[^>]+id="[^"]*(?:article|content|zoom|main-content)[^"]*"[^>]*>([\s\S]*?)</div>',
        # 次级容器
        r'<div[^>]+class="[^"]*(?:detail|main|pages_content|news_detail|txt_con)[^"]*"[^>]*>([\s\S]*?)</div>',
        # content 总容器（放在最后，容易误含导航）
        r'<div[^>]+class="[^"]*content[^"]*"[^>]*>([\s\S]*?)</div>',
    ]
    main_best = ''
    for pat in main_patterns:
        for m in re.finditer(pat, html, flags=re.I):
            txt = strip_html(m.group(1))
            if len(txt) > len(main_best):
                main_best = txt

    # 四川模板页兜底：如果 main_best 疑似只抓到导航（"当前位置"开头且字数很少），用 ywyd 块补救
    main_best_stripped = strip_html(main_best)
    if (main_best_stripped.startswith('当前位置') and len(main_best_stripped) < 300):
        # 尝试从 ywyd 块提取（贪婪匹配到外层容器倒数第二层关闭标签）
        ywyd_match = re.search(r'class="ywyd[^"]*"[^>]*>([\s\S]*?)</div>\s*</div>\s*</div>', html, flags=re.I)
        if ywyd_match:
            ywyd_block = ywyd_match.group(1)
            txt = strip_html(ywyd_block)
            if len(txt) > len(main_best_stripped):
                main_best = ywyd_block

    # Fallback pattern（JS 动态渲染页面）
    fb_patterns = [
        r'<script[^>]*>(?:window\.|var\s+)?(?:htmlContent|articleContent|content|detail)["\'].*?([\u4e00-\u9fff\u3000-\u303f\uff00-\uffef].*?)["\'];?\s*</script>',
        r'data-content="([^"]*)"',
        r'data-text="([^"]*)"',
        r'[\u4e00-\u9fff]{50,}',
    ]
    fb_best = ''
    for pat in fb_patterns:
        for m in re.finditer(pat, html, flags=re.I | re.U):
            txt = m.group(1) if m.lastindex else strip_html(m.group(0))
            txt = re.sub(r'<[^>]+>', ' ', txt)
            txt = unescape(txt).strip()
            if len(txt) > len(fb_best):
                fb_best = txt

    # 选最长的——正文肯定是最长的，无关内容很短
    return main_best if len(main_best) >= len(fb_best) else fb_best


def extract_summary(html: str, body: str) -> str:
    """提取摘要：优先 meta description，若被正文截断则用正文首段补足"""
    # 优先从 meta name=Description 拿
    desc = extract_meta(html, 'Description')
    # og:description 作为补充（只在比 Description 更长时替换）
    og_desc_m = re.search(r'<meta[^>]+property="og:description"[^>]+content="([^"]*)"', html, flags=re.I)
    if og_desc_m:
        og_desc_val = unescape(og_desc_m.group(1)).strip()
        if not desc or len(og_desc_val) > len(desc):
            desc = og_desc_val

    body_first = (body or '').strip()

    # 判断 meta 是否可用：
    # - >= 80 字：直接用（相对完整）
    # - < 80 字：检查正文是否以 meta 开头（是 → 正文截断了 meta，用正文补足；否 → 保留原 meta）
    if desc:
        if len(desc) >= 80:
            return desc
        # meta 过短：只有当 body 真正以 desc 开头（body 是 desc 的超集）才替换
        if body_first and body_first[:len(desc)].strip() == desc.strip() and len(body_first) > len(desc) + 10:
            return body_first[:200].strip()
        return desc

    # 无 meta → fallback 正文前100字
    if body_first and len(body_first) >= 100:
        return body_first[:100].strip()
    return ''


def extract_attachments(html: str) -> str:
    links = re.findall(r'<a[^>]+href="([^"]+)"[^>]*>(.*?)</a>', html, flags=re.I | re.S)
    hits = []
    for href, text in links:
        t = strip_html(text)
        if re.search(r'附件|下载|PDF|DOC|DOCX|OFD|XLS|XLSX', t, flags=re.I) or re.search(r'\.(pdf|doc|docx|ofd|xls|xlsx|zip)$', href, flags=re.I):
            hits.append(href)
    return '\n'.join(dict.fromkeys(hits))


def derive_policy_id(url: str) -> str:
    patterns = ['(t\\d+_\\d+)', '/([A-Za-z0-9_-]{16,})\\.shtml(?:\\?|$)', '/([a-f0-9]{24,})\\.shtml(?:\\?|$)', '/(\\d{10,})/', 'content[_-](\\d+)', '/info[_-](\\d+)', 'article[_-](\\d+)', 'show[_-](\\d+)', 'index[_-](\\d+)', 'page[_-](\\d+)', '[?&]id=(\\d+)', '[?&]id=(\\w+)', '[?&](?:newsId|articleId|contentId)=(\\w+)', '/(\\d{8,})/', '(\\d{8,}[-_]\\d+)', '/(\\d+)\\.html']
    for pat in patterns:
        m = re.search(pat, url, flags=re.I)
        if m:
            return m.group(1)
    return ''


def derive_region(url: str, html: str) -> str:
    region_map = {
        'hubei.gov.cn': '湖北省',
        'hunan.gov.cn': '湖南省',
        'hebei.gov.cn': '河北省',
        'yn.gov.cn': '云南省',
        'sc.gov.cn': '四川省',
        'yn.gov.cn': '云南省',
        'hunan.gov.cn': '湖南省',
    }
    for host, region in region_map.items():
        if host in url:
            return region
    site_name = extract_meta(html, 'SiteName')
    for region in ['湖北省', '湖南省', '河北省', '云南省', '四川省']:
        if region.replace('省', '') in site_name or region in site_name:
            return region
    return ''


def classify_type(title: str, summary: str, body: str) -> str:
    text = f'{title} {summary} {body}'
    rules = [
        ('项目申报', ['申报', '申报指南', '申报条件']),
        ('发展规划', ['发展规划', '规划纲要', '规划']),
        ('工作部署', ['工作要点', '工作方案', '通知', '召开', '部署']),
        ('产业扶持', ['扶持', '支持', '奖补', '补贴', '若干措施', '支持措施']),
        ('管理规范', ['管理办法', '办法', '细则', '监管']),
        ('建设实施', ['建设方案', '建设', '实施']),
        ('政策解读', ['解读', '答记者问']),
        ('统计报告', ['统计公报', '工作报告', '报告', '执行情况']),
        ('资金预算', ['预算', '资金', '财政']),
        ('人才引进', ['人才', '引进']),
    ]
    for label, keywords in rules:
        if any(k in text for k in keywords):
            return label
    return '综合管理'


def status_and_fallback(row: dict) -> tuple[str, str, str]:
    body_len = len((row.get('正文') or '').strip())
    has_attachment = '是' if (row.get('附件链接') or '').strip() else '否'
    poor = False
    if not row.get('标题'):
        poor = True
    if body_len < 120:
        poor = True
    if not row.get('发布时间'):
        poor = True
    status = 'GOOD'
    if poor and body_len > 0:
        status = 'PARTIAL'
    if poor and body_len == 0:
        status = 'POOR'
    need_browser = '是' if status in {'PARTIAL', 'POOR'} else '否'
    return status, str(body_len), need_browser if need_browser else '否'


def parse_detail(url: str) -> dict:
    row = {k: '' for k in FIELDS}
    row['原始链接'] = url
    try:
        html = fetch_text(url)
    except Exception:
        row['抓取状态'] = 'FETCH_ERROR'
        row['正文长度'] = '0'
        row['是否检测到附件'] = '否'
        row['是否需要浏览器兜底'] = '是'
        return row

    row['政策ID'] = derive_policy_id(url)
    row['标题'] = extract_title(html)
    row['正文'] = extract_body(html, url)
    row['摘要'] = extract_summary(html, row['正文'])
    row['发文机关'] = extract_meta(html, 'ContentSource')
    row['发布时间'] = extract_meta(html, 'PubDate')
    row['关键词'] = extract_meta(html, 'Keywords')
    row['附件链接'] = extract_attachments(html)
    row['地区'] = derive_region(url, html)
    row['类型'] = classify_type(row['标题'], row['摘要'], row['正文'])
    row['是否检测到附件'] = '是' if row['附件链接'] else '否'

    status, body_len, need_browser = status_and_fallback(row)
    row['抓取状态'] = status
    row['正文长度'] = body_len
    row['是否需要浏览器兜底'] = need_browser
    return row


def read_links(md_path: Path):
    text = md_path.read_text(encoding='utf-8', errors='ignore')
    urls = re.findall(r'\((http[^)]+)\)', text)
    seen = set()
    ordered = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            ordered.append(u)
    return ordered


def write_excel(rows, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = '政策详情'

    # 样式定义
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写表头
    ws.append(FIELDS)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据行
    for row_data in rows:
        ws.append([row_data.get(f, '') for f in FIELDS])
        for col_idx in range(1, len(FIELDS) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.alignment = data_align
            cell.border = thin_border

    # 设置列宽
    col_widths = {
        '标题': 40, '正文': 60, '摘要': 45,
        '发文机关': 22, '发布时间': 20,
        '原始链接': 40,
        '关键词': 22, '类型': 14, '地区': 14
    }
    for col_idx, field in enumerate(FIELDS, 1):
        width = col_widths.get(field, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 设置行高（数据行固定高度，不自动撑开）
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

    # 超链接（仅原始链接列）
    header_index = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
    if '原始链接' in header_index:
        col = header_index['原始链接']
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            value = str(cell.value or '').strip()
            if value.startswith('http'):
                cell.hyperlink = value
                cell.style = 'Hyperlink'

    wb.save(out_path)


def derive_name_from_links(link_md: Path) -> str:
    text = link_md.read_text(encoding='utf-8', errors='ignore')
    urls = re.findall(r'\((http[^)]+)\)', text)
    if not urls:
        return '搜索结果'
    host_map = {
        'www.hubei.gov.cn': '湖北省政策',
        'www.hunan.gov.cn': '湖南政策',
        'www.hebei.gov.cn': '河北政策',
        'www.yn.gov.cn': '云南省政策',
        'www.sc.gov.cn': '四川政策',
    }
    for url in urls:
        for host, name in host_map.items():
            if host in url:
                return name
    return '搜索结果'


def main():
    keyword = sys.argv[1] if len(sys.argv) > 1 else ''
    link_md = Path.home() / 'Desktop' / '搜索文件夹' / '处理文件夹' / '原文链接.md'
    out_dir = Path.home() / 'Desktop' / '搜索文件夹' / 'detail-h5'
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    keyword_name = keyword if keyword else derive_name_from_links(link_md)
    safe_keyword = re.sub(r'[\\/:*?"<>|]+', '_', keyword_name)
    out_excel = out_dir / f'{safe_keyword}_{stamp}.xlsx'

    text = link_md.read_text(encoding='utf-8', errors='ignore')
    raw_urls = re.findall(r'\((http[^)]+)\)', text)
    urls = read_links(link_md)

    def fetch_one(url):
        time.sleep(1)
        return parse_detail(url)

    with ThreadPoolExecutor(max_workers=5) as executor:
        rows = list(executor.map(fetch_one, urls))

    if out_excel.exists():
        out_excel.unlink()
    write_excel(rows, out_excel)

    print(f'INPUT_MD={link_md}')
    print(f'RAW_LINK_COUNT={len(raw_urls)}')
    print(f'UNIQUE_LINK_COUNT={len(urls)}')
    print(f'COUNT={len(rows)}')
    print(f'EXCEL={out_excel}')


if __name__ == '__main__':
    main()
