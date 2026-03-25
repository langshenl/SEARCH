#!/usr/bin/env python3
"""
政策搜索完整版 v2 - 兼顾数量与质量
- 多搜索引擎组合：Exa + Tavily + 百度
- 多关键词覆盖
- 真实正文抓取
"""
import os
import sys
import json
import time
import subprocess
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import OrderedDict

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "requests", "beautifulsoup4", "-q"])
    import requests
    from bs4 import BeautifulSoup

# 配置
EXA_API_KEY = os.environ.get("EXA_API_KEY", "25eb2029-8225-48ab-8a74-ca18f3c75987")
TAVILY_API_KEY = os.environ.get("TAVILY_API_KEY", "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")

# 省份配置
PROVINCES = OrderedDict([
    ("湖北省", "hubei.gov.cn"),
    ("北京市", "beijing.gov.cn"),
    ("广东省", "gd.gov.cn"),
    ("上海市", "shanghai.gov.cn"),
    ("浙江省", "zj.gov.cn"),
    ("江苏省", "jiangsu.gov.cn"),
    ("四川省", "sc.gov.cn"),
    ("重庆市", "cq.gov.cn"),
    ("山东省", "shandong.gov.cn"),
    ("河南省", "henan.gov.cn"),
])

# 搜索关键词组合（扩大覆盖面）
SEARCH_KEYWORDS = [
    "2026年 政策",
    "2026 通知 公告",
    "2026 规划 方案",
    "2026 工作要点",
    "2026 管理办法",
    "2026 产业发展",
    "2026 科技创新",
    "2026 民生保障",
    "2026 乡村振兴",
    "2026 绿色发展",
    "2026 数字经济",
    "2026 扩大内需",
]

def search_exa(query, domain, num_results=20):
    """使用 Exa 搜索"""
    url = "https://api.exa.ai/search"
    payload = {
        "query": f"site:{domain} {query}",
        "type": "auto",
        "numResults": num_results,
        "text": True,
        "highlights": True,
        "summary": True
    }
    headers = {
        "x-api-key": EXA_API_KEY,
        "Content-Type": "application/json"
    }
    
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        if resp.status_code == 200:
            return resp.json().get("results", [])
    except Exception as e:
        pass
    return []

def search_tavily(query, num_results=15):
    """使用 Tavily 搜索"""
    url = "https://api.tavily.com/search"
    payload = {
        "api_key": TAVILY_API_KEY,
        "query": query,
        "search_depth": "advanced",
        "max_results": num_results,
        "include_answer": False,
        "include_raw_content": True
    }
    
    try:
        resp = requests.post(url, json=payload, timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            results = data.get("results", [])
            # 转换为与 Exa 相同格式
            for r in results:
                r['url'] = r.get('url', '')
                r['title'] = r.get('title', '')
            return results
    except:
        pass
    return []

def search_baidu(query, domain, num_results=20):
    """使用百度搜索"""
    try:
        encoded_query = f"site:{domain} {query}".replace(" ", "+")
        url = f"https://www.baidu.com/s?wd={encoded_query}&rn={num_results}&ie=utf-8"
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        }
        resp = requests.get(url, headers=headers, timeout=10)
        # 简单解析百度结果
        import re
        urls = re.findall(r'href="(http[^"]*' + domain + r'[^"]*)"', resp.text)
        titles = re.findall(r'<h3[^>]*class="[^"]*t[^"]*"[^>]*>(.*?)</h3>', resp.text, re.DOTALL)
        
        results = []
        seen = set()
        for i, url in enumerate(urls[:num_results]):
            url = url.split('?')[0].split('#')[0]
            if url not in seen and domain in url:
                seen.add(url)
                title = titles[i] if i < len(titles) else ""
                title = title.replace('<em>', '').replace('</em>', '')
                results.append({
                    'url': url,
                    'title': title.strip() if title else "无标题"
                })
        return results
    except:
        return []

def fetch_content(url, timeout=10):
    """获取页面真实正文"""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        }
        resp = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
        resp.encoding = resp.apparent_encoding
        
        soup = BeautifulSoup(resp.text, 'html.parser')
        
        # 移除无用标签
        for tag in soup(['script', 'style', 'nav', 'header', 'footer', 'aside', 'iframe']):
            tag.decompose()
        
        # 查找正文区域
        article = (
            soup.find('article') or
            soup.find('div', class_=lambda x: x and ('content' in x.lower() or 'article' in x.lower())) or
            soup.find('div', id=lambda x: x and ('content' in x.lower() or 'article' in x.lower())) or
            soup.find('div', class_='main') or
            soup.find('div', class_='text') or
            soup.find('div', id='zoom')
        )
        
        if article:
            text = article.get_text(separator=' ', strip=True)
        else:
            paragraphs = soup.find_all('p')
            text = ' '.join(p.get_text(strip=True) for p in paragraphs if p.get_text(strip=True))
        
        # 清理
        text = ' '.join(text.split())
        
        # 检查是否是有用内容
        if len(text) < 100:
            return ""
        
        return text[:4000]
    except:
        return ""

def search_province_all(province_name, domain, keywords):
    """用多个关键词搜索同一省份"""
    all_results = []
    seen_urls = {}
    
    print(f"\n{'='*60}")
    print(f"🔍 搜索 {province_name} ({domain})")
    print(f"{'='*60}")
    
    total_exa = 0
    total_tavily = 0
    total_baidu = 0
    
    for kw in keywords:
        print(f"\n📌 关键词: {kw}")
        
        # 1. Exa 搜索
        try:
            results_exa = search_exa(kw, domain, num_results=15)
            total_exa += len(results_exa)
            print(f"   Exa: {len(results_exa)} 条", end="")
            for r in results_exa:
                url = r.get('url', '')
                if url and url not in seen_urls:
                    seen_urls[url] = r
        except:
            pass
        
        # 2. Tavily 搜索（带域名过滤）
        try:
            results_tavily = search_tavily(f"site:{domain} {kw}", num_results=15)
            filtered = [r for r in results_tavily if domain in r.get('url', '')]
            total_tavily += len(filtered)
            print(f" | Tavily: {len(filtered)} 条", end="")
            for r in filtered:
                url = r.get('url', '')
                if url and url not in seen_urls:
                    seen_urls[url] = r
        except:
            pass
        
        # 3. 百度搜索（补充）
        try:
            results_baidu = search_baidu(kw, domain, num_results=10)
            filtered_baidu = [r for r in results_baidu if r.get('url') and r['url'] not in seen_urls]
            total_baidu += len(filtered_baidu)
            print(f" | 百度: {len(filtered_baidu)} 条", end="")
            for r in filtered_baidu:
                seen_urls[r['url']] = r
        except:
            pass
        
        print("")
        time.sleep(0.3)  # 避免请求过快
    
    all_results = list(seen_urls.values())
    
    print(f"\n📊 {province_name} 小计:")
    print(f"   Exa: {total_exa} | Tavily: {total_tavily} | 百度: {total_baidu}")
    print(f"   去重后: {len(all_results)} 条")
    
    # 标记来源
    for r in all_results:
        r['province'] = province_name
    
    return all_results

def main():
    # 解析参数
    if len(sys.argv) < 2:
        keyword = "2026年 政策"
    else:
        keyword = sys.argv[1]
    
    # 确定搜索哪些省份
    if len(sys.argv) > 2:
        # 单省份搜索
        p_name = sys.argv[2]
        p_domain = PROVINCES.get(p_name, p_name if '.' in p_name else p_name + '.gov.cn')
        target_provinces = OrderedDict([(p_name, p_domain)])
    else:
        # 默认搜索前3个省份
        target_provinces = OrderedDict()
        for i, (p_name, p_domain) in enumerate(PROVINCES.items()):
            if i >= 3:
                break
            target_provinces[p_name] = p_domain
    
    print(f"""
╔════════════════════════════════════════════════════════════╗
║           政策搜索系统 v2 - 兼顾数量与质量                    ║
╠════════════════════════════════════════════════════════════╣
║  关键词: {keyword}
║  省份: {', '.join(target_provinces.keys())}
║  搜索轮次: {len(SEARCH_KEYWORDS)} 轮/省份
╚════════════════════════════════════════════════════════════╝
    """)
    
    all_results = []
    
    # 搜索每个省份
    for province_name, domain in target_provinces.items():
        results = search_province_all(province_name, domain, SEARCH_KEYWORDS)
        all_results.extend(results)
        time.sleep(1)
    
    print(f"\n\n{'='*60}")
    print(f"📊 总计获取: {len(all_results)} 条原始结果")
    print(f"📥 开始抓取真实正文内容...")
    print(f"{'='*60}")
    
    # 并行抓取正文
    urls_to_fetch = [(r['url'], r) for r in all_results]
    fetched_count = [0]
    failed_count = [0]
    
    def fetch_with_url(item):
        url, result = item
        content = fetch_content(url)
        if content and len(content) > 50:
            result['real_text'] = content
            result['real_summary'] = content[:400] + "..." if len(content) > 400 else content
            result['has_content'] = True
            fetched_count[0] += 1
        else:
            result['has_content'] = False
            failed_count[0] += 1
        return result
    
    with ThreadPoolExecutor(max_workers=15) as executor:
        futures = {executor.submit(fetch_with_url, item): item for item in urls_to_fetch}
        for i, future in enumerate(as_completed(futures), 1):
            if i % 20 == 0 or i == len(futures):
                print(f"   进度: {i}/{len(all_results)} | 成功: {fetched_count[0]} | 失败: {failed_count[0]}", end="\r")
            future.result()
    
    print(f"\n\n✅ 内容抓取完成!")
    print(f"   成功获取正文: {fetched_count[0]} 条")
    print(f"   获取失败: {failed_count[0]} 条")
    
    # 过滤掉没有内容的
    valid_results = [r for r in all_results if r.get('has_content', False)]
    print(f"   有效结果: {len(valid_results)} 条")
    
    # 生成输出
    output_dir = Path.home() / "Desktop" / "桌面政策文件夹"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = keyword.replace(" ", "_")[:15]
    province_str = "_".join(p[:2] for p in target_provinces.keys())
    
    # 保存JSON
    json_file = output_dir / f"政策_{province_str}_{safe_keyword}_完整版_{timestamp}.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump({
            "搜索时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "关键词": keyword,
            "省份": list(target_provinces.keys()),
            "搜索轮次": len(SEARCH_KEYWORDS),
            "原始总数": len(all_results),
            "有效总数": len(valid_results),
            "结果": valid_results
        }, f, ensure_ascii=False, indent=2)
    print(f"\n✅ JSON: {json_file}")
    
    # 生成Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"政策搜索结果"
        
        # 表头
        headers = ["序号", "标题", "正文摘要", "摘要", "来源省份", "原文地址", "发布时间"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 数据
        for idx, r in enumerate(valid_results, 2):
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=r.get('title', ''))
            ws.cell(row=idx, column=3, value=r.get('real_text', ''))
            ws.cell(row=idx, column=4, value=r.get('real_summary', ''))
            ws.cell(row=idx, column=5, value=r.get('province', ''))
            ws.cell(row=idx, column=6, value=r.get('url', ''))
            ws.cell(row=idx, column=7, value=r.get('publishedDate', '')[:10] if r.get('publishedDate') else '')
            
            for col in range(1, 8):
                ws.cell(row=idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        
        # 列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 55
        ws.column_dimensions['G'].width = 12
        
        xlsx_file = output_dir / f"政策_{province_str}_{safe_keyword}_完整版_{timestamp}.xlsx"
        wb.save(xlsx_file)
        print(f"✅ Excel: {xlsx_file}")
        
    except ImportError:
        print("⚠️ 未安装 openpyxl，仅保存JSON")
    
    # 统计
    print(f"\n{'='*60}")
    print(f"🎉 搜索完成!")
    print(f"{'='*60}")
    print(f"📊 最终结果统计:")
    print(f"   省份: {', '.join(target_provinces.keys())}")
    print(f"   关键词: {keyword}")
    print(f"   有效政策: {len(valid_results)} 条")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()