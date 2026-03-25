#!/usr/bin/env python3
"""
政策搜索完整版 - 获取真实正文内容
"""
import os
import sys
import json
import time
import subprocess
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    print("正在安装依赖...")
    subprocess.run([sys.executable, "-m", "pip", "install", "requests", "beautifulsoup4", "-q"])
    import requests
    from bs4 import BeautifulSoup

# 配置
EXA_API_KEY = os.environ.get("EXA_API_KEY", "25eb2029-8225-48ab-8a74-ca18f3c75987")
PROVINCES = {
    "湖北省": "hubei.gov.cn",
    "北京市": "beijing.gov.cn",
    "广东省": "gd.gov.cn",
    "上海市": "shanghai.gov.cn",
}

def search_exa(query, domain, num_results=30):
    """使用 Exa 搜索"""
    url = "https://api.exa.ai/search"
    payload = {
        "query": f"site:{domain} {query}",
        "type": "auto",
        "numResults": num_results,
        "text": True,
        "highlights": True,
        "summary": True
    }
    headers = {
        "x-api-key": EXA_API_KEY,
        "Content-Type": "application/json"
    }
    
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        if resp.status_code == 200:
            return resp.json().get("results", [])
    except Exception as e:
        print(f"搜索出错: {e}")
    return []

def fetch_content(url, timeout=10):
    """获取页面真实正文"""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept": "text/html,application/xhtml+xml",
        }
        resp = requests.get(url, headers=headers, timeout=timeout)
        resp.encoding = resp.apparent_encoding
        
        soup = BeautifulSoup(resp.text, 'html.parser')
        
        # 移除脚本和样式
        for tag in soup(['script', 'style', 'nav', 'header', 'footer']):
            tag.decompose()
        
        # 尝试找正文区域
        article = soup.find('article') or soup.find('div', class_='content') or soup.find('div', id='content')
        if article:
            text = article.get_text(separator=' ', strip=True)
        else:
            # 获取所有段落
            paragraphs = soup.find_all('p')
            text = ' '.join(p.get_text(strip=True) for p in paragraphs if p.get_text(strip=True))
        
        # 清理
        text = ' '.join(text.split())
        return text[:3000] if text else ""
    except Exception as e:
        return f"[获取失败: {e}]"

def search_province(province_name, domain, keyword, num_results=30):
    """搜索单个省份"""
    print(f"🔍 搜索 {province_name}...")
    
    # 多个搜索词组合
    keywords = [
        keyword,
        f"{keyword} 产业",
        f"{keyword} 民生",
        f"{keyword} 科技",
    ]
    
    all_results = []
    seen_urls = set()
    
    for kw in keywords:
        results = search_exa(kw, domain, num_results)
        for r in results:
            url = r.get('url', '')
            if url and url not in seen_urls:
                seen_urls.add(url)
                r['province'] = province_name
                all_results.append(r)
        time.sleep(0.5)  # 避免请求过快
    
    print(f"   ✅ {province_name}: 获取 {len(all_results)} 条")
    return all_results

def main():
    if len(sys.argv) < 2:
        keyword = "2026年 政策"
    else:
        keyword = sys.argv[1]
    
    province = sys.argv[2] if len(sys.argv) > 2 else "湖北省"
    domain = PROVINCES.get(province, "hubei.gov.cn")
    
    print(f"\n{'='*60}")
    print(f"政策搜索: {keyword}")
    print(f"省份: {province} ({domain})")
    print(f"{'='*60}\n")
    
    # 搜索
    results = search_province(province, domain, keyword, num_results=30)
    
    if not results:
        print("❌ 未获取到结果")
        return
    
    print(f"\n📡 获取到 {len(results)} 条原始结果")
    print("📥 开始获取真实正文内容（并行处理）...")
    
    # 并行获取正文
    urls_to_fetch = [(r['url'], r) for r in results]
    
    def fetch_with_url(item):
        url, result = item
        content = fetch_content(url)
        result['real_text'] = content
        # 生成摘要（取前300字）
        result['real_summary'] = content[:300] + "..." if len(content) > 300 else content
        return result
    
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_with_url, item): item for item in urls_to_fetch}
        for i, future in enumerate(as_completed(futures), 1):
            print(f"   处理 {i}/{len(results)}", end="\r")
            future.result()
    
    print(f"\n✅ 内容获取完成")
    
    # 生成输出文件
    output_dir = Path.home() / "Desktop" / "桌面政策文件夹"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = keyword.replace(" ", "_")[:20]
    
    # 保存JSON
    json_file = output_dir / f"{province}_{safe_keyword}_真实结果_{timestamp}.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump({
            "搜索时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "关键词": keyword,
            "省份": province,
            "总数": len(results),
            "结果": results
        }, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON: {json_file}")
    
    # 生成Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{province}政策"
        
        # 表头
        headers = ["序号", "标题", "正文摘要", "AI摘要", "来源", "原文地址", "发布时间"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # 数据
        for idx, r in enumerate(results, 2):
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=r.get('title', ''))
            ws.cell(row=idx, column=3, value=r.get('real_text', r.get('text', '')))
            ws.cell(row=idx, column=4, value=r.get('real_summary', r.get('summary', '')))
            ws.cell(row=idx, column=5, value=r.get('province', province))
            ws.cell(row=idx, column=6, value=r.get('url', ''))
            ws.cell(row=idx, column=7, value=r.get('publishedDate', '')[:10])
            
            for col in range(1, 8):
                ws.cell(row=idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        
        # 列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 12
        
        xlsx_file = output_dir / f"{province}_{safe_keyword}_真实结果_{timestamp}.xlsx"
        wb.save(xlsx_file)
        print(f"✅ Excel: {xlsx_file}")
        
    except ImportError:
        print("⚠️ 未安装 openpyxl，仅保存JSON")
    
    print(f"\n🎉 完成！共 {len(results)} 条政策")

if __name__ == "__main__":
    main()