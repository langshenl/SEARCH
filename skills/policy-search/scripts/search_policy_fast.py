#!/usr/bin/env python3
"""
政策搜索快速版 - 主题驱动查询 + 相关性过滤
- Exa 搜索
- 围绕主题本身生成查询，不再使用固定泛关键词池污染结果
- 对标题/摘要/正文做主题相关性过滤
"""
import os
import sys
import json
import time
import re
import subprocess
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import OrderedDict

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "requests", "beautifulsoup4", "-q"])
    import requests
    from bs4 import BeautifulSoup

EXA_API_KEY = os.environ.get("EXA_API_KEY", "25eb2029-8225-48ab-8a74-ca18f3c75987")

PROVINCES = OrderedDict([
    ("湖北省", "hubei.gov.cn"),
    ("北京市", "beijing.gov.cn"),
    ("广东省", "gd.gov.cn"),
    ("上海市", "shanghai.gov.cn"),
    ("浙江省", "zj.gov.cn"),
])

THEME_HINTS = {
    "博物馆": ["博物馆", "文博", "文物", "展览", "藏品", "陈列", "纪念馆", "公共文化服务", "文化遗产"],
    "农业": ["农业", "三农", "乡村振兴", "粮食", "高标准农田", "种业", "农机", "农产品"],
    "新能源": ["新能源", "光伏", "风电", "储能", "新能源汽车", "氢能", "充换电", "绿色电力"],
    "科技": ["科技", "创新", "成果转化", "技术攻关", "高新技术", "科研", "实验室"],
    "人才": ["人才", "就业", "引进", "高层次人才", "技能人才", "创业", "毕业生"],
}

GENERIC_ACTIONS = [
    "扶持政策", "发展政策", "建设政策", "管理办法", "实施方案",
    "实施意见", "若干措施", "通知", "公告", "专项规划"
]

GENERIC_POLICY_TERMS = ["政策", "通知", "公告", "方案", "意见", "办法", "措施", "规划"]


def normalize_theme(theme: str) -> str:
    theme = (theme or "").strip()
    return re.sub(r"\s+", "", theme)


def infer_core_terms(theme: str):
    core_terms = []
    for key, hints in THEME_HINTS.items():
        if key in theme:
            core_terms.extend(hints)
    if not core_terms:
        base = theme.replace("政策", "")
        base = re.sub(r"(扶持|发展|建设|管理|实施|专项|相关)$", "", base)
        if base:
            core_terms.append(base)
            core_terms.extend([f"{base}政策", f"{base}项目", f"{base}服务"])
    deduped = []
    seen = set()
    for item in core_terms:
        if item and item not in seen:
            seen.add(item)
            deduped.append(item)
    return deduped[:12]


def build_theme_queries(keyword: str):
    cleaned = normalize_theme(keyword)
    core_terms = infer_core_terms(cleaned)
    queries = [cleaned]
    for term in core_terms[:6]:
        queries.append(f"{cleaned} {term}")
    for action in GENERIC_ACTIONS[:6]:
        queries.append(f"{cleaned} {action}")
    deduped = []
    seen = set()
    for q in queries:
        q = re.sub(r"\s+", " ", q).strip()
        if q and q not in seen:
            seen.add(q)
            deduped.append(q)
    return deduped[:12], core_terms


def search_exa(query, domain, num_results=20):
    url = "https://api.exa.ai/search"
    payload = {
        "query": f"site:{domain} {query}",
        "type": "auto",
        "numResults": num_results,
        "text": True,
        "highlights": True,
        "summary": True,
    }
    headers = {
        "x-api-key": EXA_API_KEY,
        "Content-Type": "application/json",
    }
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        if resp.status_code == 200:
            return resp.json().get("results", [])
    except Exception:
        pass
    return []


def fetch_content(url, timeout=10):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept-Language": "zh-CN,zh;q=0.9",
        }
        resp = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
        resp.encoding = resp.apparent_encoding
        soup = BeautifulSoup(resp.text, 'html.parser')
        for tag in soup(['script', 'style', 'nav', 'header', 'footer']):
            tag.decompose()
        article = (
            soup.find('article') or
            soup.find('div', class_=lambda x: x and 'content' in str(x).lower()) or
            soup.find('div', id=lambda x: x and 'content' in str(x).lower()) or
            soup.find('div', class_='main') or
            soup.find('div', class_='text')
        )
        if article:
            text = article.get_text(separator=' ', strip=True)
        else:
            paragraphs = soup.find_all('p')
            text = ' '.join(p.get_text(strip=True) for p in paragraphs if p.get_text(strip=True))
        text = ' '.join(text.split())
        if len(text) < 100:
            return ""
        return text[:4000]
    except Exception:
        return ""


def relevance_score(text: str, core_terms):
    text = text or ""
    score = 0
    for term in core_terms:
        if term and term in text:
            score += 1
    return score


def is_relevant_result(result, core_terms):
    title = result.get('title', '') or ''
    summary = result.get('summary', '') or ''
    highlights = ' '.join(result.get('highlights', []) or [])
    text = f"{title} {summary} {highlights}"
    score = relevance_score(text, core_terms)
    if score >= 1:
        return True
    if any(term in title for term in core_terms[:3]):
        return True
    return False


def is_relevant_content(result, core_terms):
    text = f"{result.get('title', '')} {result.get('real_text', '')} {result.get('real_summary', '')}"
    score = relevance_score(text, core_terms)
    if score >= 2:
        return True
    if any(term in text for term in core_terms[:2]):
        return True
    return False


def main():
    keyword = sys.argv[1] if len(sys.argv) > 1 else "2026年 政策"
    debug_dir = Path.home() / "Desktop" / "搜索配置文件夹"
    debug_dir.mkdir(parents=True, exist_ok=True)
    debug_log = debug_dir / "exa-search-debug.log"
    with debug_log.open('a', encoding='utf-8') as f:
        f.write(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] keyword={keyword}\n")

    if len(sys.argv) > 2:
        p_name = sys.argv[2]
        p_domain = PROVINCES.get(p_name, p_name + '.gov.cn')
        target_provinces = OrderedDict([(p_name, p_domain)])
    else:
        target_provinces = OrderedDict()
        for i, (p_name, p_domain) in enumerate(PROVINCES.items()):
            if i >= 3:
                break
            target_provinces[p_name] = p_domain

    query_variants, core_terms = build_theme_queries(keyword)
    with debug_log.open('a', encoding='utf-8') as f:
        f.write(f"  core_terms={core_terms}\n")
        for q in query_variants:
            f.write(f"  query_variant={q}\n")

    print(f"""
╔════════════════════════════════════════════════════════════╗
║           政策搜索系统 v3 - 主题驱动 + 相关性过滤             ║
╠════════════════════════════════════════════════════════════╣
║  关键词: {keyword}
║  省份: {', '.join(target_provinces.keys())}
║  搜索轮次: {len(query_variants)} 轮/省份
╚════════════════════════════════════════════════════════════╝
    """)

    all_results = []
    for province_name, domain in target_provinces.items():
        print(f"\n🔍 搜索 {province_name} ({domain})...")
        seen_urls = {}
        for query in query_variants:
            with debug_log.open('a', encoding='utf-8') as f:
                f.write(f"  province={province_name} domain={domain} actual_query={query}\n")
            results = search_exa(query, domain, num_results=15)
            filtered = [r for r in results if is_relevant_result(r, core_terms)]
            with debug_log.open('a', encoding='utf-8') as f:
                f.write(f"    raw_count={len(results)} filtered_count={len(filtered)}\n")
            for r in filtered:
                url = r.get('url', '')
                if url and url not in seen_urls:
                    seen_urls[url] = r
            time.sleep(0.2)
        results_list = list(seen_urls.values())
        for r in results_list:
            r['province'] = province_name
        print(f"   ✅ 获取 {len(results_list)} 条初筛结果")
        all_results.extend(results_list)

    print(f"\n📊 总计获取: {len(all_results)} 条初筛结果")
    print("📥 抓取真实正文内容...")

    urls_to_fetch = [(r['url'], r) for r in all_results]
    fetched = [0]
    failed = [0]

    def fetch_with_url(item):
        url, result = item
        content = fetch_content(url)
        if content and len(content) > 50:
            result['real_text'] = content
            result['real_summary'] = content[:400] + "..." if len(content) > 400 else content
            result['has_content'] = True
            fetched[0] += 1
        else:
            result['has_content'] = False
            failed[0] += 1
        return result

    with ThreadPoolExecutor(max_workers=20) as executor:
        futures = {executor.submit(fetch_with_url, item): item for item in urls_to_fetch}
        for i, future in enumerate(as_completed(futures), 1):
            if i % 20 == 0 or i == len(futures):
                print(f"   进度: {i}/{len(all_results)} | 成功: {fetched[0]} | 失败: {failed[0]}", end="\r")
            future.result()

    valid_results = [r for r in all_results if r.get('has_content', False)]
    relevant_results = [r for r in valid_results if is_relevant_content(r, core_terms)]

    print(f"\n\n✅ 完成! 初筛有效结果: {len(valid_results)} 条")
    print(f"🎯 主题过滤后结果: {len(relevant_results)} 条")

    output_dir = Path.home() / "Desktop" / "桌面政策文件夹"
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = re.sub(r"[^\u4e00-\u9fa5A-Za-z0-9_]+", "_", keyword)[:30]
    province_str = "_".join(p[:2] for p in target_provinces.keys())

    json_file = output_dir / f"政策_{province_str}_{safe_keyword}_完整版_{timestamp}.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump({
            "搜索时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "关键词": keyword,
            "核心词": core_terms,
            "查询变体": query_variants,
            "省份": list(target_provinces.keys()),
            "原始总数": len(all_results),
            "有效总数": len(valid_results),
            "主题过滤后总数": len(relevant_results),
            "结果": relevant_results
        }, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON: {json_file}")

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "政策搜索结果"
        headers = ["序号", "标题", "正文摘要", "摘要", "来源省份", "原文地址", "发布时间"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for idx, r in enumerate(relevant_results, 2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=r.get('title', ''))
            ws.cell(row=idx, column=3, value=r.get('real_text', ''))
            ws.cell(row=idx, column=4, value=r.get('real_summary', ''))
            ws.cell(row=idx, column=5, value=r.get('province', ''))
            ws.cell(row=idx, column=6, value=r.get('url', ''))
            ws.cell(row=idx, column=7, value=r.get('publishedDate', '')[:10] if r.get('publishedDate') else '')
            for col in range(1, 8):
                ws.cell(row=idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 55
        ws.column_dimensions['G'].width = 12
        xlsx_file = output_dir / f"政策_{province_str}_{safe_keyword}_完整版_{timestamp}.xlsx"
        wb.save(xlsx_file)
        print(f"✅ Excel: {xlsx_file}")
    except Exception as e:
        print(f"⚠️ Excel生成失败: {e}")

    print(f"\n🎉 搜索完成! 共 {len(relevant_results)} 条主题相关政策")

if __name__ == "__main__":
    main()
