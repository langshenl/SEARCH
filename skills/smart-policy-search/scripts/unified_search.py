#!/usr/bin/env python3
"""
智能政策搜索工具
- CDP爬虫：用于已知官网，深度爬取
- Exa API：用于全网搜索、跨省份

用法:
  python3 unified_search.py <关键词> [--source cdp|exa|auto] [--output dir]
  python3 unified_search.py "社会力量" --source auto
  python3 unified_search.py "湖北文旅政策" --source cdp
  python3 unified_search.py "全国新能源政策" --source exa

自动模式(source=auto):
  - 检测到指定官网(文旅部、湖北等) → CDP爬虫
  - 全网搜索(全国、新能源等) → Exa API
  - 多省份 → Exa循环搜索各省
"""

import argparse
import atexit
import json
import os
import re
import signal
import subprocess
import sys
import time
import urllib.parse
from concurrent.futures import ThreadPoolExecutor

from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print('missing dependency: openpyxl', file=sys.stderr)
    sys.exit(2)

# 导入官网配置
from gov_config import (
    PROVINCE_KEYWORDS,
    DEPT_CONFIGS, DEPT_KEYWORDS, detect_dept_source,
    GENERIC_LINK_SELECTORS, GENERIC_CONTENT_SELECTORS
)

# ============================================================
# CDP爬虫相关（从mct_search.py适配）
# ============================================================

CDP = "http://localhost:3456"
CONCURRENCY = 8
SEARCH_WAIT = 4
ARTICLE_WAIT = 1.0
MAX_RETRIES = 1

# 全局tab跟踪，用于异常时清理
_open_tabs = set()

# 百度搜索配置
BAIDU_SEARCH_URL = "https://www.baidu.com/s?wd={keyword}"
BAIDU_LINK_SELECTORS = [
    "h3.t a",
    ".result a",
    ".c-container a",
    "h3 a",
]
BAIDU_ABSTRACT_SELECTORS = [
    ".c-abstract",
    ".abstract",
    ".c-summary",
    ".result-op .c-abstract",
]
BAIDU_CONTENT_SELECTORS = [
    ".TRS_Editor",
    ".article-content",
    ".article",
    ".detail-content",
    ".content",
    "#zoom",
    "article",
    "main",
    ".main-content",
    "[class*='content']",
    "[id*='content']",
    ".bd-article",
    ".article-body",
    ".post-content",
    "#article",
]

GENERIC_BAIDU_CONFIG = {
    "link_selectors": BAIDU_LINK_SELECTORS,
    "content_selectors": BAIDU_CONTENT_SELECTORS,
    "support_tables": False,
}


def cdp_new(url):
    for attempt in range(MAX_RETRIES + 1):
        try:
            resp = subprocess.run(
                ['curl', '-s', f'{CDP}/new?url={url}'],
                capture_output=True, text=True, timeout=25
            )
            target_id = json.loads(resp.stdout.strip()).get('targetId', '')
            if target_id:
                _open_tabs.add(target_id)
                return target_id
        except Exception as e:
            if attempt == MAX_RETRIES:
                print(f"  [警告] cdp_new 失败: {e}")
        if attempt < MAX_RETRIES:
            time.sleep(1)
    return ''


def cdp_eval(target_id, script, timeout=15):
    for attempt in range(MAX_RETRIES + 1):
        try:
            resp = subprocess.run(
                ['curl', '-s', '-X', 'POST', f'{CDP}/eval?target={target_id}', '-d', '@-'],
                input=script.encode('utf-8'), capture_output=True, timeout=timeout
            )
            stdout = resp.stdout.strip()
            obj = json.loads(stdout)
            return obj.get('value', '')
        except Exception as e:
            if attempt == MAX_RETRIES:
                print(f"  [警告] cdp_eval 失败: {e}")
        if attempt < MAX_RETRIES:
            time.sleep(1)
    return ''


def cdp_close(target_id):
    try:
        subprocess.run(['curl', '-s', f'{CDP}/close?target={target_id}'], timeout=10)
        _open_tabs.discard(target_id)
    except Exception:
        pass


def cleanup_all_tabs():
    """关闭所有打开的tab，用于异常清理"""
    for target_id in list(_open_tabs):
        cdp_close(target_id)
    _open_tabs.clear()


# 信号处理：被 kill 前清理残留 tab
def _signal_cleanup(signum, frame):
    cleanup_all_tabs()
    sys.exit(128 + signum)

signal.signal(signal.SIGTERM, _signal_cleanup)
signal.signal(signal.SIGINT, _signal_cleanup)
atexit.register(cleanup_all_tabs)


def clean_text(text):
    """清洗正文中的导航噪声和页脚版权信息"""
    if not text:
        return ''

    # 页脚版权噪音（各部门精确匹配，避免误删正文）
    text = re.sub(r'版权所有：中华人民共和国文化和旅游部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国教育部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国财政部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国公安部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国民政部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国商务部[^\n]*', '', text)
    text = re.sub(r'版权所有：人力资源和社会保障部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国自然资源部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国生态环境部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国住房和城乡建设部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国交通运输部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国水利部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国农业农村部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国卫生健康委员会[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国应急管理部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国科学技术部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国工业和信息化部[^\n]*', '', text)
    text = re.sub(r'版权所有：中华人民共和国司法部[^\n]*', '', text)
    text = re.sub(r'版权所有：国家市场监督管理总局[^\n]*', '', text)
    text = re.sub(r'版权所有：国家体育总局[^\n]*', '', text)
    text = re.sub(r'版权所有：国家统计局[^\n]*', '', text)
    text = re.sub(r'版权所有：国家知识产权局[^\n]*', '', text)
    text = re.sub(r'版权所有：国家烟草专卖局[^\n]*', '', text)
    text = re.sub(r'版权所有：国家铁路局[^\n]*', '', text)
    text = re.sub(r'版权所有：中国民用航空局[^\n]*', '', text)
    text = re.sub(r'版权所有：国家邮政局[^\n]*', '', text)
    text = re.sub(r'版权所有：国家中医药管理局[^\n]*', '', text)
    text = re.sub(r'版权所有：国务院[^\n]*', '', text)
    # 通用页脚噪音
    text = re.sub(r'网站地图\s*版权保护\s*免责声明\s*关于我们', '', text)
    text = re.sub(r'地址：[^/\n]{5,50}', '', text)
    text = re.sub(r'邮编：\d+', '', text)
    text = re.sub(r'ICP备案：[^\s]+', '', text)
    text = re.sub(r'网站标识码[^\s]*', '', text)
    text = re.sub(r'京公网安备[^\s]+', '', text)
    text = re.sub(r'电话：\d+', '', text)

    # 导航噪音
    text = re.sub(r'分享到：[^\n]*\n', '', text)
    text = re.sub(r'分享到：', '', text)
    text = re.sub(r'【返回顶部】【打印本页】【关闭窗口】', '', text)
    text = re.sub(r'上[^\n]*?篇：[^\n]*', '', text)
    text = re.sub(r'下[^\n]*?篇：[^\n]*', '', text)
    text = re.sub(r'按回车键[^\n]*导盲模式。', '', text)
    text = re.sub(r'来源：[^\n]*', '', text)
    text = re.sub(r'编辑：[^\n]*', '', text)
    text = re.sub(r'时间：[^\n]*', '', text)
    text = re.sub(r'发布日期：[^\n]*', '', text)
    text = re.sub(r'首\s*页[^\n]*', '', text)
    text = re.sub(r'工作邮箱[^\n]*', '', text)
    text = re.sub(r'文旅办公\s*\|\s*简\s*\|\s*繁\s*\|\s*无障碍浏览.*?(?=机构简介|$)', '', text, flags=re.DOTALL)
    text = re.sub(r'机构简介\s*信息发布\s*政务公开\s*政务服务\s*公共服务\s*互动交流', '', text)
    text = re.sub(r'当前位置：.*?(?=\n|$)', '', text)
    text = re.sub(r'^\d+\s+$', '', text, flags=re.MULTILINE)
    text = re.sub(r'\n{3,}', '\n\n', text)

    return text.strip()


def wait_for_page_load(target_id, timeout=10):
    """轮询等待页面加载完成，检测页面内容长度"""
    for i in range(timeout):
        script = "document.body ? document.body.innerText.length : 0"
        result = cdp_eval(target_id, script)
        try:
            body_len = int(result) if result else 0
        except Exception:
            body_len = 0
        if body_len > 500:
            return True
        time.sleep(0.5)
    return False


DOWNLOAD_EXTS = {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.zip', '.rar', '.7z', '.tar', '.gz', '.csv', '.txt', '.rtf', '.wps', '.et', '.dps'}

def _is_download_link(url):
    """判断URL是否为下载链接（文件扩展名或含download关键词）"""
    try:
        path = urllib.parse.urlparse(url).path.lower()
        if any(path.endswith(ext) for ext in DOWNLOAD_EXTS):
            return True
        if 'download' in path or '/attachment/' in path or '/upload/' in path:
            return True
    except Exception:
        pass
    return False


def extract_links(target_id, config):
    """从搜索页提取链接（过滤下载链接）"""
    selectors = config.get("link_selectors", []) or GENERIC_LINK_SELECTORS
    domain = config.get("domain", "")
    all_links = []

    for selector in selectors:
        script = f"""var r=[];var els=document.querySelectorAll('{selector}');els.forEach(function(a){{if(a.href)r.push({{text:a.textContent.replace(/<[^>]+>/g,'').trim(),href:a.href}})}});JSON.stringify(r)"""
        result = cdp_eval(target_id, script)
        if result:
            try:
                links = json.loads(result)
                if links:
                    all_links.extend(links)
            except Exception:
                pass

    # 回退：通用选择器无有效结果时，用域名匹配提取有效链接（过滤导航）
    has_valid = any(len(l.get('text', '')) > 15 for l in all_links)
    if (not all_links or not has_valid) and domain:
        fallback_script = f"""var r=[];var els=document.querySelectorAll('a');els.forEach(function(a){{var t=a.textContent.replace(/<[^>]+>/g,'').trim();if(a.href && t.length>15 && a.href.indexOf('{domain}')!==-1)r.push({{text:t,href:a.href}})}});JSON.stringify(r)"""
        result = cdp_eval(target_id, fallback_script)
        if result:
            try:
                all_links = json.loads(result) or []
            except Exception:
                pass

    # 去重 + 过滤下载链接
    seen = set()
    unique_links = []
    for link in all_links:
        href = link['href']
        if href not in seen and link['text'].strip():
            if _is_download_link(href):
                continue
            seen.add(href)
            unique_links.append(link)

    return unique_links


# 翻页选择器（按优先级）
PAGE_NEXT_SELECTORS = [
    "a.next",
    "a.page-next",
    ".next a",
    ".page-item.next a",
    "li.next a",
    "a[rel='next']",
    "a[title='下一页']",
    "button.next",
    ".pagination .next",
    ".pager .next",
]


def cdp_click_next(target_id):
    """点击下一页，返回是否成功"""
    # 方式1：遍历所有链接找文字包含"下一页"的并点击
    script1 = """var result='none';var btns=document.querySelectorAll('a, button');for(var i=0;i<btns.length;i++){var t=btns[i].textContent.trim();if(t==='下一页'||t==='下页'||t==='下一頁'){btns[i].click();result='ok';break;}}result;"""
    result = cdp_eval(target_id, script1)
    if result == 'ok':
        return True

    # 方式2：用选择器直接点击
    for selector in PAGE_NEXT_SELECTORS:
        script2 = f"""var el=document.querySelector('{selector}');if(el){{el.click();'ok'}}else{{'none'}}"""
        result2 = cdp_eval(target_id, script2)
        if result2 == 'ok':
            return True

    return False


def extract_article_text(target_id, config):
    """提取文章正文，支持表格"""
    selectors = config.get("content_selectors", []) or GENERIC_CONTENT_SELECTORS
    support_tables = config.get("support_tables", False)

    for selector in selectors:
        # 检查是否包含表格
        if support_tables:
            check_script = f"""var el=document.querySelector('{selector}');if(el){{var t=el.innerHTML||'';JSON.stringify({{hasTable:t.indexOf('<table')!==-1}})}}else{{'null'}}"""
            check_result = cdp_eval(target_id, check_script)
            has_table = False
            if check_result and check_result != 'null':
                try:
                    check_data = json.loads(check_result)
                    has_table = check_data.get('hasTable', False)
                except Exception:
                    pass

            if has_table:
                # 提取表格数据
                table_script = f"""var result=[];var el=document.querySelector('{selector}');if(el){{var tables=el.querySelectorAll('table');tables.forEach(function(tbl){{var rows=tbl.querySelectorAll('tr');var data=[];rows.forEach(function(row){{var cells=row.querySelectorAll('td, th');var rowData=[];cells.forEach(function(c){{rowData.push(c.textContent.replace(/\\s+/g,' ').trim())}});if(rowData.length>0)data.push(rowData)}});if(data.length>0)result.push(data)}});JSON.stringify(result)}}else{{'null'}}"""
                table_result = cdp_eval(target_id, table_script)
                if table_result and table_result != 'null':
                    try:
                        table_data = json.loads(table_result)
                        if table_data and len(table_data) > 0:
                            text_script = f"""var el=document.querySelector('{selector}');if(el){{var t=el.innerText.trim();JSON.stringify(t)}}else{{'null'}}"""
                            text_result = cdp_eval(target_id, text_script)
                            text = clean_text(json.loads(text_result)) if text_result and text_result != 'null' else ''
                            return {'selector': selector, 'text': text, 'tables': table_data}
                    except Exception:
                        pass

        # 无表格，正常提取（阈值降到30，避免短正文容器被跳过）
        script = f"""var el=document.querySelector('{selector}');if(el){{var t=el.innerText.trim();if(t.length>30){{JSON.stringify({{selector:'{selector}',text:t}})}}else{{'null'}}}}else{{'null'}}"""
        result = cdp_eval(target_id, script)
        if result and result != 'null':
            try:
                data = json.loads(result)
                if data and data.get('text'):
                    data['text'] = clean_text(data['text'])
                    data['tables'] = []
                    return data
            except Exception:
                pass

    # 回退
    raw = cdp_eval(target_id, 'document.body.innerText')
    return {'selector': 'body(fallback)', 'text': clean_text(raw) if raw else '', 'tables': []}


def _fetch_articles(urls, config):
    """批量抓取文章正文，返回 {url: {text, tables, selector}}，失败自动重试1次"""
    results = {}
    total = len(urls)
    failed_urls = []  # 收集失败URL用于重试

    def _do_batch(batch_urls, batch_offset, is_retry=False):
        nonlocal results
        bt = len(batch_urls)
        for batch_start in range(0, bt, CONCURRENCY):
            batch = batch_urls[batch_start:batch_start + CONCURRENCY]
            batch_num = batch_start // CONCURRENCY + 1
            total_batches = (bt + CONCURRENCY - 1) // CONCURRENCY
            prefix = "  [重试]" if is_retry else ""
            print(f"{prefix}  批次 {batch_num}/{total_batches}...")
            with ThreadPoolExecutor(max_workers=CONCURRENCY) as executor:
                targets = list(executor.map(cdp_new, batch))
            try:
                def _wait_ready(t):
                    if not t:
                        return
                    for _ in range(3):
                        try:
                            state_info = subprocess.run(['curl', '-s', f'{CDP}/info?target={t}'], capture_output=True, text=True, timeout=10)
                            state = json.loads(state_info.stdout.strip()).get('ready', '')
                            if state in ('complete', 'interactive'):
                                return
                        except Exception:
                            pass
                        time.sleep(0.5)
                with ThreadPoolExecutor(max_workers=CONCURRENCY) as executor:
                    list(executor.map(_wait_ready, targets))
                time.sleep(ARTICLE_WAIT)
                def _extract_one(item):
                    i, url, t = item
                    if not t:
                        return None
                    try:
                        info_resp = subprocess.run(['curl', '-s', f'{CDP}/info?target={t}'], capture_output=True, text=True, timeout=10)
                        resolved_url = json.loads(info_resp.stdout.strip()).get('url', '') or ''
                    except Exception:
                        resolved_url = ''
                    try:
                        data = extract_article_text(t, config)
                    except Exception:
                        data = None
                    text = data.get('text', '') if data else ''
                    tables = data.get('tables', []) if data else []
                    selector = data.get('selector', 'unknown')[:20] if data else 'unknown'
                    cdp_close(t)
                    targets[i] = None
                    return {'i': i, 'url': url, 'text': text, 'tables': tables, 'selector': selector, 'real_url': resolved_url}
                items = [(i, url, t) for i, (url, t) in enumerate(zip(batch, targets))]
                with ThreadPoolExecutor(max_workers=CONCURRENCY) as executor:
                    extract_results = list(executor.map(_extract_one, items))
                for r in extract_results:
                    if not r:
                        continue
                    idx = batch_offset + batch_start + r['i'] + 1
                    has_table = " [含表格]" if r['tables'] else ""
                    if r['text'] and len(r['text']) > 50:
                        results[r['url']] = {'text': r['text'], 'tables': r['tables'], 'selector': r['selector'], 'real_url': r['real_url']}
                        print(f"    ✓ [{idx}/{total}] ({r['selector']}, {len(r['text'])}字){has_table}")
                    else:
                        print(f"    ✗ [{idx}/{total}] 内容为空")
                        if not is_retry:
                            failed_urls.append(r['url'])
            except Exception:
                for t in targets:
                    if t:
                        cdp_close(t)
                raise
            if batch_start + CONCURRENCY < bt:
                time.sleep(0.5)

    # 第一轮爬取
    _do_batch(urls, 0)

    # 重试失败的URL（仅1次）
    if failed_urls:
        retry_urls = list(set(failed_urls))
        print(f"\n  重试 {len(retry_urls)} 篇失败文章...")
        failed_urls.clear()
        _do_batch(retry_urls, 0, is_retry=True)

    return results


def cdp_dept_search(keyword, dept_name, output_dir=None, max_results=30):
    """使用CDP爬虫搜索部门官网（DEPT_CONFIGS格式），按总条数翻页"""
    if dept_name not in DEPT_CONFIGS:
        print(f"[ERROR] 未知部门: {dept_name}")
        return None

    config = DEPT_CONFIGS[dept_name]
    domain = config["domain"]
    gov_name = dept_name
    search_url_tpl = config.get("search_url", "")

    if not search_url_tpl:
        print(f"[跳过CDP] {dept_name}（{domain}）搜索URL未验证，跳过CDP")
        return None

    search_url = search_url_tpl.format(keyword=urllib.parse.quote(keyword))

    print(f"\n{'='*50}")
    print(f"[CDP爬虫] {gov_name}")
    print(f"域名: {domain}")
    print(f"关键词: {keyword}")
    print(f"搜索页: {search_url}")
    print(f"目标: 最多 {max_results} 条结果")
    print(f"{'='*50}")

    # 构造兼容配置（统一用GENERIC_SELECTORS，support_tables从配置读取）
    dept_config = {
        "name": gov_name,
        "domain": domain,
        "link_selectors": GENERIC_LINK_SELECTORS,
        "content_selectors": GENERIC_CONTENT_SELECTORS,
        "support_tables": config.get("support_tables", False),
    }

    # Step 1: 打开搜索页
    print("\n[1/4] 打开搜索页...")
    search_mode = config.get("search_mode", "url")
    search_target = cdp_new(search_url)
    if not search_target:
        print("ERROR: 无法打开搜索页，请确认CDP服务运行中")
        return None
    wait_for_page_load(search_target)

    # 表单提交模式：输入关键词并点击搜索
    if search_mode == "form":
        input_sel = config.get("search_input", "")
        button_sel = config.get("search_button", "")
        if input_sel and button_sel:
            print(f"  表单提交模式: 输入关键词 → 点击搜索")
            safe_kw = json.dumps(keyword)
            safe_input = json.dumps(input_sel)
            safe_button = json.dumps(button_sel)
            form_script = f"""var el=document.querySelector({safe_input});if(el){{el.value={safe_kw};el.dispatchEvent(new Event('input',{{bubbles:true}}));var btn=document.querySelector({safe_button});if(btn)btn.click();'ok'}}else{{'no_input'}}"""
            cdp_eval(search_target, form_script)
            time.sleep(3)
            wait_for_page_load(search_target)

    # Step 2: 提取链接（支持翻页，去重）
    print("[2/4] 提取链接（翻页中）...")
    all_links = []
    seen = set()
    page = 0
    max_pages = 3  # 最多翻3页
    stop_reason = ''
    while True:
        page += 1
        links = extract_links(search_target, dept_config)
        for l in links:
            if l['href'] not in seen and l['text'].strip():
                seen.add(l['href'])
                all_links.append(l)
        print(f"  第 {page} 页: 本页 {len(links)} 条, 累计 {len(all_links)} 条")

        if len(all_links) >= max_results:
            all_links = all_links[:max_results]
            stop_reason = f"已达到目标 {max_results} 条"
            break

        if page >= max_pages:
            stop_reason = f"已达最大 {max_pages} 页"
            break

        clicked = cdp_click_next(search_target)
        if not clicked:
            stop_reason = "无下一页"
            break
        wait_for_page_load(search_target)

    cdp_close(search_target)
    print(f"  {stop_reason}，停止翻页")
    print(f"  共提取到 {len(all_links)} 条链接")

    if not all_links:
        print("  未提取到任何链接，跳过")
        return None

    for i, l in enumerate(all_links[:10]):
        print(f"  {i+1:3d}. {l['text'][:40]}")

    # Step 3: 爬取正文
    print(f"\n[3/4] 爬取正文 ({len(all_links)} 篇，{CONCURRENCY}并发)...")
    urls = [l['href'] for l in all_links]
    results = _fetch_articles(urls, dept_config)
    print(f"  成功爬取 {len(results)} / {len(urls)} 篇")

    # Step 4: 生成Excel
    if output_dir is None:
        output_dir = Path("~/Desktop/smart搜索文件夹").expanduser()
    excel_path = generate_excel(all_links, results, keyword, gov_name, output_dir)
    return excel_path


def extract_baidu_abstracts(target_id):
    """从百度搜索页提取标题+摘要+时间，返回 [{title, href, abstract, time}] """
    # 新百度页面结构：h3.t > a 是标题，cos-row 是摘要（在同一title-wrapper的兄弟元素里）
    # 时间通常在 c-color-gray 类的 span 中
    script = """
var results = [];
var h3s = document.querySelectorAll('h3.t');
for (var i = 0; i < h3s.length; i++) {
    var h3 = h3s[i];
    var titleEl = h3.querySelector('a');
    if (!titleEl || !titleEl.href) continue;
    var titleWrapper = h3.parentElement ? h3.parentElement.parentElement : null;
    var abstractText = '';
    var timeText = '';
    if (titleWrapper && titleWrapper.parentElement && titleWrapper.parentElement.children.length > 1) {
        var cosRow = titleWrapper.parentElement.children[1];
        if (cosRow && cosRow.className.indexOf('cos-row') !== -1) {
            abstractText = cosRow.textContent.replace(/\s+/g, ' ').trim();
        }
    }
    // 提取时间：通常在标题附近的 c-color-gray 或 c-font-small 元素中
    var timeEl = h3.parentElement ? h3.parentElement.querySelector('.c-color-gray, .c-font-small, .c-color-gray2') : null;
    if (timeEl) {
        timeText = timeEl.textContent.replace(/\s+/g, ' ').trim();
    }
    results.push({
        text: titleEl.textContent.replace(/<[^>]+>/g, '').trim(),
        href: titleEl.href,
        abstract: abstractText,
        time: timeText
    });
}
results.length > 0 ? JSON.stringify(results) : JSON.stringify([{error: 'no results'}]);
"""
    result = cdp_eval(target_id, script)
    if not result:
        return []
    try:
        data = json.loads(result)
        if isinstance(data, list) and len(data) > 0 and 'error' in data[0]:
            return []
        return data
    except Exception:
        return []


def cdp_baidu_search(keyword, max_results=10):
    """
    使用CDP爬百度，取前max_results条结果，优先用百度摘要，终端汇总输出。
    失败后返回None（由调用方决定是否降级Exa）。
    """
    search_url = BAIDU_SEARCH_URL.format(keyword=urllib.parse.quote(keyword))

    print(f"\n{'='*50}")
    print(f"[CDP百度搜索] {keyword}")
    print(f"搜索页: {search_url}")
    print(f"{'='*50}")

    # Step 1: 打开百度搜索页
    print("\n[1/3] 打开百度搜索页...")
    search_target = cdp_new(search_url)
    if not search_target:
        print("ERROR: 无法打开百度搜索页，请确认CDP服务运行中")
        return None
    wait_for_page_load(search_target)

    # Step 2: 提取链接+摘要
    print("[2/3] 提取搜索结果...")
    raw_results = extract_baidu_abstracts(search_target)
    cdp_close(search_target)

    if not raw_results:
        print("  未提取到任何搜索结果")
        return None

    # 去重（按原始 baidu 跳转 URL 去重）
    seen = set()
    unique_results = []
    for r in raw_results:
        if not r['href'] or not r['text'].strip():
            continue
        if r['href'] not in seen:
            seen.add(r['href'])
            unique_results.append(r)
    unique_results = unique_results[:max_results]

    print(f"  获取到 {len(unique_results)} 条结果")
    if not unique_results:
        print("  未提取到任何链接")
        return None

    for i, r in enumerate(unique_results, 1):
        print(f"  {i}. {r['text'][:50]}")
        if r['abstract']:
            print(f"     摘要: {r['abstract'][:80]}...")

    # Step 3: 汇总输出（优先用百度摘要，全文爬取作为补充）
    # 直接用原始 baidu 跳转 URL 爬取（CDP 会自动跟随跳转）
    # 注意：resolve_baidu_url 对 baidu 链接返回的是 base64 参数值（无效 URL），
    #       故改用原始 href，CDP /new 会正确解析并导航到真实页面
    url_to_info = {r['href']: (r['text'], r['abstract'], r.get('time', '')) for r in unique_results}
    crawl_urls = list(url_to_info.keys())

    # 尝试爬取全文（作为摘要的补充）
    print(f"\n[3/3] 尝试补充爬取全文 ({len(crawl_urls)} 篇)...")
    articles = _fetch_articles(crawl_urls, GENERIC_BAIDU_CONFIG)
    fetched_ok = len(articles)
    print(f"  成功爬取 {fetched_ok} / {len(crawl_urls)} 篇")

    # 汇总输出（只输出有实质内容的来源，跳过抓取失败的）
    valid_count = 0
    print(f"\n{'='*60}")
    print("【汇总分析】")
    print(f"{'='*60}")
    print(f"针对您的问题「{keyword}」，从 {len(crawl_urls)} 篇搜索结果中汇总如下：\n")

    for url, (title, abstract_text, time_text) in url_to_info.items():
        content = articles.get(url, {}).get('text', '')

        # 跳过无实质内容的来源（无摘要 + 爬取失败）
        if not abstract_text and not content:
            continue

        valid_count += 1
        # 优先用百度摘要（有值就直接用，不设长度门槛）
        # 仅在完全没有摘要时才尝试正文内容
        if abstract_text:
            display_text = abstract_text
        elif content:
            display_text = content[:300]
        else:
            display_text = ""

        print(f"--- 来源{valid_count} ---")
        print(f"标题: {title}")
        if time_text:
            print(f"时间: {time_text}")
        print(f"摘要: {display_text[:300]}...")
        if articles.get(url):
            real_link = articles[url].get('real_url', url)
            print(f"原始链接: {real_link}")
        print()

    return True  # 完成，返回True表示成功（终端输出，不生成Excel）


EXA_API_KEY = os.environ.get("EXA_API_KEY", "")


def exa_search(keyword, region=None, output_dir=None):
    """使用Exa API搜索"""
    if not EXA_API_KEY:
        print("ERROR: EXA_API_KEY未设置，请先设置环境变量:")
        print('  export EXA_API_KEY="your-exa-api-key"')
        print("获取地址: https://exa.ai")
        return None, None

    # 构建查询
    site_filter = None
    if region:
        # region 有值 → 加site过滤
        # "全国" → site:gov.cn，其他省份 → site:xxx.gov.cn
        query = keyword
        site_filter = f"site:{region_to_domain(region)}"
        full_query = f"{query} {site_filter}"
    else:
        # region is None → 不加过滤，全网搜
        query = keyword
        full_query = keyword

    print(f"\n{'='*50}")
    print(f"[Exa API] 全网搜索")
    print(f"关键词: {query}")
    if site_filter:
        print(f"站点过滤: {site_filter}")
    print(f"{'='*50}")

    # 调用Exa API
    curl_cmd = [
        'curl', '-s', '-X', 'POST', 'https://api.exa.ai/search',
        '-H', 'Content-Type: application/json',
        '-H', f'Authorization: Bearer {EXA_API_KEY}',
        '-H', 'User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        '-d', json.dumps({
            "query": full_query,
            "numResults": 30,
            "type": "auto",
            "contents": {"text": True, "summary": True}
        })
    ]

    print("\n[1/3] 调用Exa API...")
    try:
        result = subprocess.run(curl_cmd, capture_output=True, text=True, timeout=90)
        data = json.loads(result.stdout)
        items = data.get('results', [])
        print(f"  获取到 {len(items)} 条结果")
    except Exception as e:
        raise RuntimeError(f"Exa API调用失败: {e}")

    if not items:
        raise RuntimeError("Exa API未返回任何结果")

    # 处理结果
    print("\n[2/3] 处理结果...")
    rows = []
    for item in items:
        url = item.get('url', '')
        title = item.get('title', '')
        text = clean_text(item.get('text', ''))
        summary = item.get('summary', '')[:120]
        pub_date = item.get('publishedDate', '')

        if len(text) < 50:
            continue

        rows.append({
            '标题': title or url,
            '正文': text,
            '摘要': summary,
            '发布时间': pub_date[:10] if pub_date else '',
            '原始链接': url,
            '关键词': keyword,
            '类型': infer_type(keyword, text),
            '地区': region or '全国',
        })

    print(f"  有效结果: {len(rows)} 条")

    if not rows:
        print("  无有效结果，跳过生成Excel")
        return rows, None

    # 生成Excel
    if output_dir is None:
        output_dir = Path("~/Desktop/smart搜索文件夹").expanduser()

    print("\n[3/3] 生成Excel...")
    excel_path = generate_excel_simple(rows, keyword, region or '全国', output_dir)

    return rows, excel_path


def region_to_domain(region):
    """省份转域名；如果region本身是域名（包含.）则直接返回"""
    if not region:
        return "gov.cn"
    # 如果包含点，当作域名直接返回
    if '.' in region:
        return region
    domain_map = {
        "北京": "beijing.gov.cn",
        "天津": "tj.gov.cn",
        "河北": "hebei.gov.cn",
        "山西": "shanxi.gov.cn",
        "内蒙古": "nmg.gov.cn",
        "辽宁": "ln.gov.cn",
        "吉林": "jl.gov.cn",
        "黑龙江": "hlj.gov.cn",
        "上海": "shanghai.gov.cn",
        "江苏": "jiangsu.gov.cn",
        "浙江": "zj.gov.cn",
        "安徽": "ah.gov.cn",
        "福建": "fujian.gov.cn",
        "江西": "jiangxi.gov.cn",
        "山东": "shandong.gov.cn",
        "河南": "henan.gov.cn",
        "湖北": "hubei.gov.cn",
        "湖南": "hunan.gov.cn",
        "广东": "gd.gov.cn",
        "广西": "gx.gov.cn",
        "海南": "hainan.gov.cn",
        "重庆": "cq.gov.cn",
        "四川": "sc.gov.cn",
        "贵州": "gz.gov.cn",
        "云南": "yn.gov.cn",
        "西藏": "xz.gov.cn",
        "陕西": "shaanxi.gov.cn",
        "甘肃": "gs.gov.cn",
        "青海": "qh.gov.cn",
        "宁夏": "nx.gov.cn",
        "新疆": "xj.gov.cn",
        "全国": "gov.cn",
    }
    return domain_map.get(region, "gov.cn")


def infer_type(keyword, text):
    """简单推断文档类型"""
    combined = keyword + text
    rules = [
        ('项目申报', ['申报', '征集', '遴选']),
        ('政策解读', ['解读', '答记者问']),
        ('发展规划', ['规划', '计划', '纲要']),
        ('工作部署', ['通知', '工作要点', '部署']),
        ('产业扶持', ['扶持', '补贴', '奖补', '支持']),
    ]
    for doc_type, keywords in rules:
        if any(kw in combined for kw in keywords):
            return doc_type
    return '综合管理'


# ============================================================
# Excel生成
# ============================================================

COLUMNS = ["标题", "正文", "摘要", "发布时间", "原始链接", "关键词", "类型", "地区"]
COL_WIDTHS = {'标题': 30, '正文': 50, '摘要': 30, '发布时间': 15, '原始链接': 60, '关键词': 20, '类型': 12, '地区': 12}


def _setup_worksheet(ws):
    """设置表头样式"""
    ws.append(COLUMNS)
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 15)
    ws.row_dimensions[1].height = 22


def generate_excel_simple(rows, keyword, region, output_dir):
    """生成Excel（用于Exa结果，rows为字典列表）"""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    region_prefix = f"{region}_" if region != '全国' else ""
    excel_path = output_dir / f"{region_prefix}{keyword}_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = '搜索结果'
    _setup_worksheet(ws)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if col_name == '原始链接':
                url = row.get(col_name, '')
                if url:
                    cell.hyperlink = url
                    cell.value = url
                    cell.font = Font(color='0563C1', underline='single')
            else:
                cell.value = row.get(col_name, '')
        ws.row_dimensions[row_idx].height = 60

    wb.save(excel_path)
    return excel_path


def generate_excel(links, results, keyword, source, output_dir):
    """生成Excel（用于CDP结果，links为链接列表，results为正文字典）"""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = output_dir / f"{source}_{keyword}_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = '搜索结果'
    _setup_worksheet(ws)

    row_idx = 2

    for link in links:
        url = link['href']
        title = link['text']
        data = results.get(url, {})
        text = data.get('text', '')
        tables = data.get('tables', [])

        if len(text) < 50:
            continue

        body = f"【含{len(tables)}个表格，内容从略】\n\n{text[:8000]}" if tables else text[:32000]

        ws.cell(row=row_idx, column=1, value=title)
        ws.cell(row=row_idx, column=2, value=body)
        ws.cell(row=row_idx, column=3, value=text[:200] + "..." if len(text) > 200 else text)
        ws.cell(row=row_idx, column=4, value='')
        cell_link = ws.cell(row=row_idx, column=5, value=url)
        cell_link.hyperlink = url
        cell_link.font = Font(color='0563C1', underline='single')
        ws.cell(row=row_idx, column=6, value=keyword)
        ws.cell(row=row_idx, column=7, value='政策文件')
        ws.cell(row=row_idx, column=8, value=source)

        for col_idx in range(1, 9):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[row_idx].height = 80
        row_idx += 1

    wb.save(excel_path)
    return excel_path


# ============================================================
# 统一调度
# ============================================================

def extract_search_keyword(query):
    """
    从查询中提取核心搜索词

    例如: "文旅部关于社会力量的政策" → "社会力量"
          "湖北省文旅政策" → "文旅"
          "碳中和是什么" → "碳中和"
    """
    keyword = query

    # 去掉省份名
    for p in PROVINCE_KEYWORDS:
        keyword = keyword.replace(p, " ")

    # 去掉"省"字
    keyword = keyword.replace("省", " ")

    # 去掉部门关键词简称（DEPT_KEYWORDS的所有key）
    for kw, dept_name in DEPT_KEYWORDS.items():
        keyword = keyword.replace(kw, " ")

    # 去掉"官网"、"官方网站"等词
    keyword = keyword.replace("官网", " ")
    keyword = keyword.replace("官方网站", " ")

    # 去掉常见修饰词（保留政策类意图词）
    noise_words = [
        "关于", "相关",
        "工作", "活动", "项目", "建设", "实施",
        "印发", "发布", "转发",
        "如何", "怎么", "怎么办", "是什么", "为什么",
        "好不好", "怎么样", "多少", "哪里", "哪个",
        "介绍一下", "解释一下", "说一说", "讲讲",
        "什么", "搜索",
    ]
    # 政策类意图词：原查询包含则保留
    policy_intent_words = ["政策", "通知", "公告", "规定", "办法", "条例"]
    for w in noise_words:
        keyword = keyword.replace(w, " ")

    # 保留原查询中的政策意图词
    for w in policy_intent_words:
        if w in query and w not in keyword:
            keyword = keyword + " " + w

    # 去掉多余空格
    keyword = " ".join(keyword.split())

    # 如果去完后为空或太短，用原查询
    if len(keyword.strip()) < 2:
        keyword = query

    return keyword.strip()


def auto_decide_source(query):
    """
    路由规则（按优先级）:
    1. 含"官网" + 多部门 → cdp_multi（每个部门一个Excel）
    2. 含"官网" + 单部门 → cdp
    3. 含"官网" + 无映射 → exa（site:gov.cn）
    4. 日常问题 → cdp_baidu（失败→Exa全网降级）
    5. 省份 + 政策类关键词 → exa（site:省.gov.cn，搜不到就不出）
    6. 有省份但非政策类 → exa（site:省.gov.cn，搜不到就不出）
    7. 默认/泛搜索 → cdp_baidu（失败→Exa全网降级）

    返回: (source_mode, dept_names, provinces, search_kw)
      - cdp_multi时 dept_names 为部门名列表
      - cdp时 dept_names 为单一部门名列表
      - exa时 dept_names 为空列表，provinces 为省份列表
    """
    # 1. 含"官网" → 部门检测
    gov_keywords = ["官网", "官方网站"]
    if any(kw in query for kw in gov_keywords):
        dept_names = detect_dept_source(query)
        if dept_names:
            # 多部门 → cdp_multi，每个部门单独爬
            if len(dept_names) > 1:
                return "cdp_multi", dept_names, [], extract_search_keyword(query)
            # 单部门 → cdp
            return "cdp", dept_names, [], extract_search_keyword(query)
        # 无部门匹配 → Exa兜底
        return "exa", [], [], query

    # 2. 检测省份
    detected_provinces = []
    for p in PROVINCE_KEYWORDS:
        if p in query:
            detected_provinces.append(p)
    has_provinces = len(detected_provinces) > 0

    # 4. 省份 + 政策/机构类关键词 → Exa（site:省.gov.cn，失败→CDP百度降级）
    policy_keywords = ["政策", "通知", "公告", "规定", "办法", "条例",
                      "博物馆", "图书馆", "文化馆", "美术馆", "纪念馆",
                      "文旅", "文化", "旅游", "体育", "教育"]
    if has_provinces and any(kw in query for kw in policy_keywords):
        return "exa", [], detected_provinces, extract_search_keyword(query)

    # 4. 有省份但非政策类 → Exa（site:省.gov.cn）
    if has_provinces:
        return "exa", [], detected_provinces, extract_search_keyword(query)

    # 5. 默认 → Exa全网搜索
    return "exa", [], [], query


def unified_search(query, source="auto", output_dir=None, max_results=30):
    """统一搜索入口"""
    if output_dir is None:
        output_dir = Path("~/Desktop/smart搜索文件夹").expanduser()

    print(f"\n{'='*60}")
    print(f"统一政策搜索")
    print(f"查询: {query}")
    print(f"{'='*60}")

    # 自动决策
    if source == "auto":
        source_mode, dept_names, provinces, search_kw = auto_decide_source(query)
        print(f"[自动模式] 选择: {source_mode}")
        if dept_names:
            print(f"[部门] {dept_names}")
        if provinces:
            print(f"[省份] {provinces}")
        print(f"[搜索词] {search_kw}")
    else:
        source_mode = source
        dept_names = []
        provinces = []
        search_kw = query  # 非auto模式用原始查询

    excel_paths = []

    try:
        if source_mode == "cdp":
            dept_name = dept_names[0] if dept_names else None
            if dept_name:
                excel_path = cdp_dept_search(search_kw, dept_name, output_dir, max_results)
                if excel_path:
                    excel_paths.append(excel_path)
                else:
                    # 部门CDP跳过或失败 → 用Exa搜该部门域名
                    dept_domain = DEPT_CONFIGS.get(dept_name, {}).get("domain", "")
                    if dept_domain:
                        result = exa_search(search_kw, dept_domain, output_dir)
                    else:
                        result = exa_search(search_kw, None, output_dir)
                    if result:
                        rows, excel_path = result
                        if excel_path:
                            excel_paths.append(excel_path)
            else:
                result = exa_search(search_kw, None, output_dir)
                if result:
                    _, excel_path = result
                    if excel_path:
                        excel_paths.append(excel_path)

        elif source_mode == "cdp_multi":
            # 多部门搜索，并行爬取
            print(f"[多部门搜索] 共 {len(dept_names)} 个部门: {dept_names}")

            def _search_dept(dept_name):
                print(f"\n  → 搜索 {dept_name}: {search_kw}")
                excel_path = cdp_dept_search(search_kw, dept_name, output_dir, max_results)
                if excel_path:
                    return excel_path
                # CDP失败 → 降级Exa
                dept_domain = DEPT_CONFIGS.get(dept_name, {}).get("domain", "")
                fallback_region = dept_domain if dept_domain else None
                print(f"    {dept_name} CDP失败，Exa降级搜索...")
                result = exa_search(search_kw, fallback_region, output_dir)
                if result:
                    _, ep = result
                    return ep
                return None

            with ThreadPoolExecutor(max_workers=len(dept_names)) as executor:
                results = list(executor.map(_search_dept, dept_names))
            excel_paths.extend([p for p in results if p])

        elif source_mode == "exa":
            if provinces:
                # 多省：每省用 Exa（site:省.gov.cn），搜不到就不出
                print(f"[多省搜索] 共 {len(provinces)} 个省份: {provinces}")
                for prov in provinces:
                    print(f"\n  → {prov}: {search_kw}")
                    result = exa_search(search_kw, prov, output_dir)
                    if result:
                        _, excel_path = result
                        if excel_path:
                            excel_paths.append(excel_path)
                    else:
                        print(f"    {prov} 未搜索到结果，跳过")
            else:
                # 无省份：Exa 全网搜索，直接输出汇总
                rows, excel_path = exa_search(search_kw, None, output_dir)
                if rows:
                    print(f"\n{'='*60}")
                    print("【总结分析】")
                    print(f"{'='*60}")
                    print(f"针对您的问题「{search_kw}」，搜索到 {len(rows)} 条参考信息：\n")
                    for i, r in enumerate(rows[:5], 1):
                        title = r.get('标题', '无标题')[:60]
                        print(f"{i}. {title}")
                    if len(rows) > 5:
                        print(f"\n...还有其他 {len(rows)-5} 条结果")
                    if rows[0].get('摘要'):
                        print(f"\n【参考】")
                        print(f"{rows[0]['摘要'][:300]}...")

        else:
            print(f"ERROR: 未知source: {source}")

    except Exception as e:
        print(f"\n⚠️ 搜索出错: {e}")
        if source_mode in ("cdp", "cdp_multi", "cdp_baidu"):
            print("CDP爬虫失败，切换到 Exa API 保底...")
            try:
                if provinces:
                    for prov in provinces:
                        result = exa_search(search_kw, prov, output_dir)
                        if result:
                            _, excel_path = result
                            if excel_path:
                                excel_paths.append(excel_path)
                elif dept_names:
                    for dept_name in dept_names:
                        dept_domain = DEPT_CONFIGS.get(dept_name, {}).get("domain", "")
                        result = exa_search(search_kw, dept_domain or None, output_dir)
                        if result:
                            rows, excel_path = result
                            if excel_path:
                                excel_paths.append(excel_path)
                else:
                    result = exa_search(search_kw, None, output_dir)
                    if result:
                        rows, excel_path = result
                        if rows:
                            print(f"降级成功，获取到 {len(rows)} 条结果")
            except Exception as ex2:
                print(f"❌ 保底方案也失败了: {ex2}")
                print("请检查网络连接或稍后重试")
        else:
            print(f"❌ Exa API 失败: {e}")
            print("请检查 EXA_API_KEY 是否正确，或稍后重试")

    # 输出结果
    print(f"\n{'='*60}")
    print("搜索完成!")
    if excel_paths:
        print("生成文件:")
        for ep in excel_paths:
            print(f"  📁 {ep}")
    else:
        print("未生成任何文件")

    return excel_paths


# ============================================================
# 命令行入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description='智能政策搜索工具（CDP爬虫 + Exa API）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  python3 unified_search.py "社会力量"                    # 自动选择数据源
  python3 unified_search.py "文旅部社会力量" --source cdp    # 强制CDP
  python3 unified_search.py "全国新能源政策" --source exa     # 强制Exa
  python3 unified_search.py "湖北文旅政策" --source exa       # 省份+政策→Exa
  python3 unified_search.py "文旅部官网搜索社会力量"           # 官网→CDP部门爬虫
  python3 unified_search.py "全国新能源政策" --output ~/Desktop  # 指定输出目录
  python3 unified_search.py "通知" --max-results 100          # CDP爬虫最多抓100条

注意: --max-results 仅对CDP爬虫有效，Exa API固定返回30条
        '''
    )
    parser.add_argument('keyword', nargs='?', help='搜索关键词')
    parser.add_argument('--source', '-s', choices=['auto', 'cdp', 'exa'], default='auto',
                        help='数据源: auto自动判断, cdp强制CDP爬虫, exa强制Exa API')
    parser.add_argument('--output', '-o', default=None,
                        help='输出目录，默认 ~/Desktop/smart搜索文件夹')
    parser.add_argument('--max-results', '-n', type=int, default=30,
                        help='CDP爬虫最大条数，默认30（仅对CDP有效，Exa固定30条）')

    args = parser.parse_args()

    if not args.keyword:
        parser.print_help()
        print("\n请提供搜索关键词")
        sys.exit(1)

    try:
        unified_search(args.keyword, source=args.source, output_dir=args.output, max_results=args.max_results)
    finally:
        cleanup_all_tabs()


if __name__ == '__main__':
    main()
