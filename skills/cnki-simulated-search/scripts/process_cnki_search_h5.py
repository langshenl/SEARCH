#!/usr/bin/env python3
from __future__ import annotations

import re
import sys
from html import unescape
from pathlib import Path
from urllib.parse import urljoin
from openpyxl import Workbook

BASE = 'https://kns.cnki.net'


def strip_html(text: str) -> str:
    text = re.sub(r'<script[\s\S]*?</script>', ' ', text, flags=re.I)
    text = re.sub(r'<style[\s\S]*?</style>', ' ', text, flags=re.I)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = unescape(text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def extract_results(html: str):
    results = []
    patterns = [
        re.compile(r'<a[^>]+href="([^"]+)"[^>]*target="_blank"[^>]*>([\s\S]*?)</a>', re.I),
    ]
    for pat in patterns:
        for m in pat.finditer(html):
            href = unescape(m.group(1)).strip()
            title = strip_html(m.group(2))
            if not title or len(title) < 2:
                continue
            full = urljoin(BASE, href)
            if 'kns.cnki.net' not in full and '/kcms/' not in full and '/kns8s/' not in full:
                continue
            results.append({
                '标题': title,
                '作者': '',
                '来源': '',
                '年份': '',
                '详情链接': full,
            })
    dedup = []
    seen = set()
    for r in results:
        key = (r['标题'], r['详情链接'])
        if key in seen:
            continue
        seen.add(key)
        dedup.append(r)
    return dedup


def write_outputs(rows, out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    excel = out_dir / 'cnki_搜索结果.xlsx'
    md = out_dir / 'cnki_详情链接.md'

    wb = Workbook()
    ws = wb.active
    ws.title = '搜索结果'
    ws.append(['标题', '作者', '来源', '年份', '详情链接'])
    for row in rows:
        ws.append([row['标题'], row['作者'], row['来源'], row['年份'], row['详情链接']])
    idx = 5
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=idx)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.style = 'Hyperlink'
    wb.save(excel)

    lines = ['# CNKI 详情链接', '']
    for i, row in enumerate(rows, 1):
        lines.append(f"{i}. [{row['标题']}]({row['详情链接']})")
    lines.append('')
    md.write_text('\n'.join(lines), encoding='utf-8')
    return excel, md


def main():
    if len(sys.argv) < 2:
        print('用法: python3 process_cnki_search_h5.py <h5文件路径>')
        return 1
    path = Path(sys.argv[1]).expanduser()
    html = path.read_text(encoding='utf-8', errors='ignore')
    rows = extract_results(html)
    out_dir = Path.home() / 'Desktop' / 'cnki-output'
    excel, md = write_outputs(rows, out_dir)
    print(f'INPUT: {path}')
    print(f'COUNT: {len(rows)}')
    print(f'EXCEL: {excel}')
    print(f'MD: {md}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
