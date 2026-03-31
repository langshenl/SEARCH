#!/usr/bin/env python3
from __future__ import annotations

import re
import urllib.request
from html import unescape
from pathlib import Path
from openpyxl import Workbook

USER_AGENT = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36'
TIMEOUT = 20
FIELDS = ['标题', '摘要', '关键词', '作者', '作者机构', '来源', '文献类型', '详情链接', '抓取状态']


def fetch(url: str) -> str:
    req = urllib.request.Request(url, headers={'User-Agent': USER_AGENT, 'Accept-Language': 'zh-CN,zh;q=0.9'})
    with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
        charset = resp.headers.get_content_charset() or 'utf-8'
        return resp.read().decode(charset, errors='ignore')


def strip_html(text: str) -> str:
    text = re.sub(r'<script[\s\S]*?</script>', ' ', text, flags=re.I)
    text = re.sub(r'<style[\s\S]*?</style>', ' ', text, flags=re.I)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = unescape(text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def extract_meta_block(html: str, label: str) -> str:
    m = re.search(rf'{re.escape(label)}[：:]?</span>([\s\S]*?)</p>', html, flags=re.I)
    return strip_html(m.group(1)) if m else ''


def parse_detail(url: str) -> dict:
    row = {k: '' for k in FIELDS}
    row['详情链接'] = url
    try:
        html = fetch(url)
    except Exception:
        row['抓取状态'] = 'FETCH_ERROR'
        return row
    title_m = re.search(r'<title>([\s\S]*?)</title>', html, flags=re.I)
    row['标题'] = strip_html(title_m.group(1)) if title_m else ''
    row['摘要'] = extract_meta_block(html, '摘要')
    row['关键词'] = extract_meta_block(html, '关键词')
    row['作者'] = extract_meta_block(html, '作者')
    row['作者机构'] = extract_meta_block(html, '作者单位')
    row['来源'] = extract_meta_block(html, '来源')
    row['文献类型'] = extract_meta_block(html, '分类号') or extract_meta_block(html, '专题')
    row['抓取状态'] = 'GOOD' if row['标题'] or row['摘要'] else 'PARTIAL'
    return row


def read_links(md_path: Path):
    text = md_path.read_text(encoding='utf-8', errors='ignore')
    urls = re.findall(r'\((http[^)]+)\)', text)
    seen = set()
    ordered = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            ordered.append(u)
    return ordered


def write_excel(rows, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'CNKI详情'
    ws.append(FIELDS)
    for row in rows:
        ws.append([row.get(f, '') for f in FIELDS])
    idx = FIELDS.index('详情链接') + 1
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=idx)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.style = 'Hyperlink'
    wb.save(out_path)


def main():
    md_path = Path.home() / 'Desktop' / 'cnki-output' / 'cnki_详情链接.md'
    out_path = Path.home() / 'Desktop' / 'cnki-output' / 'cnki_详情结果.xlsx'
    urls = read_links(md_path)
    rows = [parse_detail(u) for u in urls]
    write_excel(rows, out_path)
    print(f'INPUT_MD={md_path}')
    print(f'COUNT={len(rows)}')
    print(f'EXCEL={out_path}')


if __name__ == '__main__':
    main()
