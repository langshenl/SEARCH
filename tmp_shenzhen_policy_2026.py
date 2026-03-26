import requests,re,os,json,time
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADERS={'User-Agent':'Mozilla/5.0'}
BASE='/Users/chenyiming/.openclaw/workspace-search'
out_dir=os.path.expanduser('~/Desktop/政策文件夹')
os.makedirs(out_dir, exist_ok=True)

policy_words=['通知','公告','实施意见','工作方案','管理办法','实施细则','指南','通告','规定','计划','若干措施','申报指南','办事指南','扶持办法','奖励办法','行动方案','操作规程','实施办法']

def get(url):
    r=requests.get(url,headers=HEADERS,timeout=20)
    r.raise_for_status()
    return r.text

def clean(s):
    return re.sub(r'\s+',' ',s or '').strip()

def extract_date(text):
    m=re.search(r'(20\d{2})[-年/\.](\d{1,2})[-月/\.](\d{1,2})',text)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return ''

def summarize_text(text, maxlen=140):
    text=clean(text)
    for bad in ['无障碍浏览','繁體版','English','移动版','数据开放','政务邮箱','站群导航','政务机器人']:
        text=text.replace(bad,'')
    return text[:maxlen]

def fetch_page_meta(url):
    html=get(url)
    soup=BeautifulSoup(html,'html.parser')
    title=''
    if soup.title:
        title=clean(soup.title.get_text())
        title=title.split('--')[0].split('_')[0].strip()
    if not title:
        h=soup.find(['h1','h2'])
        title=clean(h.get_text()) if h else ''
    text=clean(soup.get_text(' ',strip=True))
    date=extract_date(text)
    # summary: prefer article paragraphs
    paras=[]
    for sel in ['.TRS_Editor p','.view p','.article p','.content p','.txt p','p']:
        found=soup.select(sel)
        if found:
            paras=[clean(p.get_text(' ',strip=True)) for p in found if clean(p.get_text(' ',strip=True))]
            if paras:
                break
    summary=''
    for p in paras:
        if len(p)>=20 and not re.match(r'^(附件|扫一扫|分享到|打印|关闭)',p):
            summary=p
            break
    if not summary:
        summary=summarize_text(text, 160)
    return title,date,summary,text,soup

formal=[]
interpret=[]
seen_urls=set()
seen_formal_titles=set()

# 1) crawl official interpretation pages and reverse to original docs
for i in range(1,5):
    idx='index.html' if i==1 else f'index_{i}.html'
    url=f'https://www.sz.gov.cn/zfgb/zcjd/{idx}'
    html=get(url)
    soup=BeautifulSoup(html,'html.parser')
    for li in soup.find_all('li'):
        text=clean(li.get_text(' ',strip=True))
        if '2026年' not in text and '2026-' not in text:
            continue
        a=li.find('a',href=True)
        if not a:
            continue
        interp_url=urljoin(url,a['href'])
        if interp_url in seen_urls:
            continue
        seen_urls.add(interp_url)
        interp_title=clean(a.get('title') or a.get_text())
        interp_date=extract_date(text) or '发布时间待二次确认'
        try:
            ititle,idate,isummary,itext,isoup=fetch_page_meta(interp_url)
            if ititle: interp_title=ititle
            if idate: interp_date=idate
            origin_url=''
            origin_title=''
            # pick substantive link in article body
            for aa in isoup.find_all('a',href=True):
                t=clean(aa.get_text(' ',strip=True))
                href=urljoin(interp_url,aa['href'])
                if href==interp_url: continue
                if href.startswith('javascript:'): continue
                if ('政策解读' in t and '/zcjd/' in href):
                    continue
                if any(k in t for k in ['通知','办法','方案','措施','实施意见','细则','规定','通告','指南','公告','计划','操作规程']) or '/zfgb/2026/' in href or '/gkmlpt/content/' in href:
                    origin_url=href
                    origin_title=t
                    break
            interpret.append({
                '标题':interp_title,
                '摘要':isummary,
                '来源地方':'深圳',
                '来源网站':re.sub(r'^https?://','',interp_url).split('/')[0],
                '原文地址':interp_url,
                '发布时间':interp_date,
                '结果类型':'官方解读',
                '地域':'深圳',
                '年份':'2026',
                '主题':'全部政策',
                '文种/页面类型':'政策解读',
                '可信度评分':88,
                '备注':'政府公报政策解读页',
            })
            if origin_url:
                ftitle,fdate,fsummary,ftext,fsoup=fetch_page_meta(origin_url)
                title=ftitle or origin_title or interp_title.replace('政策解读','').strip('《》')
                note='原始政策源（由政策解读页反向定位）'
                if not fdate:
                    fdate='发布时间待二次确认'
                    note+='；发布时间待二次确认'
                if '2026' not in fdate and fdate!='发布时间待二次确认':
                    continue
                if title in seen_formal_titles:
                    continue
                seen_formal_titles.add(title)
                formal.append({
                    '标题':title,
                    '摘要':fsummary,
                    '来源地方':'深圳',
                    '来源网站':re.sub(r'^https?://','',origin_url).split('/')[0],
                    '原文地址':origin_url,
                    '发布时间':fdate,
                    '结果类型':'原始政策',
                    '地域':'深圳',
                    '年份':'2026',
                    '主题':'全部政策',
                    '文种/页面类型':next((w for w in policy_words if w in title), '政策正文'),
                    '可信度评分':95,
                    '备注':note,
                })
        except Exception as e:
            print('interp fail',interp_url,e)

# 2) supplement official dept/district execution guides and notices
supplement_urls=[
 'https://fgw.sz.gov.cn/gkmlpt/content/12/12696/post_12696731.html',
 'https://pnr.sz.gov.cn/szsghhzrzyjwzgkml/szsghhzrzyjwzgkml/ghjh/csgh/zxgh/content/post_12687158.html',
 'https://stic.sz.gov.cn/xxgk/tzgg/content/post_12681398.html',
 'https://stic.sz.gov.cn/xxgk/tzgg/content/post_12686508.html',
 'https://stic.sz.gov.cn/xxgk/tzgg/content/post_12690655.html',
 'https://stic.sz.gov.cn/hdjlpt/yjzj/answer/49390',
 'https://stic.sz.gov.cn/xxgk/tzgg/content/post_12697182.html',
 'https://stic.sz.gov.cn/xxgk/tzgg/content/post_12697181.html',
 'https://stic.sz.gov.cn/xxgk/tzgg/content/post_12697180.html',
 'https://stic.sz.gov.cn/xxgk/tzgg/content/post_12697179.html',
 'https://www.szns.gov.cn/xxgk/qzfxxgkml/tzgg/content/post_12694405.html',
 'https://www.lg.gov.cn/bmzz/rlzyj/xxgk/qt/tzgg/content/post_12702163.html'
]
for url in supplement_urls:
    try:
        title,date,summary,text,soup=fetch_page_meta(url)
        if not title:
            continue
        if '深圳' not in text and '南山区' not in text and '龙岗区' not in text:
            continue
        date = date or '发布时间待二次确认'
        note='市级/区级执行类通知、申报指南或征求意见'
        score=90 if 'sz.gov.cn' in url else 86
        if title in seen_formal_titles:
            continue
        if not any(k in title for k in policy_words) and not any(k in text[:500] for k in ['申报指南','管理办法','工作方案','征求意见','实施','政策']):
            continue
        seen_formal_titles.add(title)
        formal.append({
            '标题':title,
            '摘要':summary,
            '来源地方':'深圳',
            '来源网站':re.sub(r'^https?://','',url).split('/')[0],
            '原文地址':url,
            '发布时间':date,
            '结果类型':'原始政策',
            '地域':'深圳',
            '年份':'2026',
            '主题':'全部政策',
            '文种/页面类型':next((w for w in policy_words if w in title), '执行通知/指南'),
            '可信度评分':score,
            '备注':note + ('；发布时间待二次确认' if date=='发布时间待二次确认' else ''),
        })
    except Exception as e:
        print('supp fail',url,e)

# filter region/time and dedupe by url/title
uniq=[]
seen=set()
for row in formal:
    key=(row['标题'],row['原文地址'])
    if key in seen: continue
    seen.add(key)
    if row['发布时间']!='发布时间待二次确认' and not row['发布时间'].startswith('2026-'):
        continue
    uniq.append(row)
formal=sorted(uniq,key=lambda x:(x['发布时间']=='发布时间待二次确认',x['发布时间']), reverse=True)

uniq=[]
seen=set()
for row in interpret:
    key=(row['标题'],row['原文地址'])
    if key in seen: continue
    seen.add(key)
    if row['发布时间']!='发布时间待二次确认' and not row['发布时间'].startswith('2026-'):
        continue
    uniq.append(row)
interpret=sorted(uniq,key=lambda x:(x['发布时间']=='发布时间待二次确认',x['发布时间']), reverse=True)

headers=['标题','摘要','来源地方','来源网站','原文地址','发布时间','结果类型','地域','年份','主题','文种/页面类型','可信度评分','备注']

def write_xlsx(path, rows):
    wb=Workbook()
    ws=wb.active
    ws.title='Sheet1'
    ws.append(headers)
    fill=PatternFill('solid', fgColor='D9EAF7')
    for c in ws[1]:
        c.font=Font(bold=True)
        c.fill=fill
        c.alignment=Alignment(vertical='center', wrap_text=True)
    for row in rows:
        ws.append([row.get(h,'') for h in headers])
    widths={'A':45,'B':60,'C':10,'D':24,'E':52,'F':14,'G':12,'H':10,'I':8,'J':10,'K':18,'L':10,'M':28}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment=Alignment(vertical='top', wrap_text=True)
    wb.save(path)

formal_path=os.path.join(out_dir,'深圳_2026_全部政策_正式政策.xlsx')
interp_path=os.path.join(out_dir,'深圳_2026_全部政策_官方解读转载.xlsx')
write_xlsx(formal_path, formal)
write_xlsx(interp_path, interpret)

print(json.dumps({
    'formal_count':len(formal),
    'interpret_count':len(interpret),
    'formal_path':formal_path,
    'interpret_path':interp_path,
    'sample_formal':[r['标题'] for r in formal[:8]],
    'sample_interpret':[r['标题'] for r in interpret[:5]],
}, ensure_ascii=False, indent=2))
