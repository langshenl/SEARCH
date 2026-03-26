import re, json, time, urllib.parse, os, sys
from pathlib import Path
from collections import OrderedDict, defaultdict
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

UA='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36'
HEADERS={'User-Agent':UA,'Accept-Language':'zh-CN,zh;q=0.9'}
SESSION=requests.Session()
SESSION.headers.update(HEADERS)

REGION='广东'
YEAR='2026'
THEME='全部政策'

DOC_TERMS=['通知','公告','实施意见','工作方案','若干措施','管理办法','实施细则','指南','规定','计划']
THEMES=['财政','科技','产业','创新','企业','中小企业','制造业','数字经济','人才','教育','医疗','乡村振兴','农业','消费','投资','招商','外贸','环保','能源','新能源汽车']
DEPTS=['发改委','财政厅','科技厅','工信厅','教育厅','人社厅','商务厅','农业农村厅','卫健委','住建厅','文旅厅','市场监管局','数据局','生态环境厅']
PROV_DOMAIN='gd.gov.cn'
CITY_DOMAINS=['gz.gov.cn','sz.gov.cn','foshan.gov.cn','dg.gov.cn','zhuhai.gov.cn','shantou.gov.cn','jiangmen.gov.cn','huizhou.gov.cn','zhaoqing.gov.cn','qingyuan.gov.cn','maoming.gov.cn','zhanjiang.gov.cn','heyuan.gov.cn','meizhou.gov.cn','shanwei.gov.cn','yangjiang.gov.cn','chaozhou.gov.cn','yunfu.gov.cn','shaoguan.gov.cn','jieyang.gov.cn']
A_MEDIA_HINTS=['人民网','新华网','南方网','羊城晚报','南方日报']

BAD_TITLE_HINTS=['考试','招聘','录用','成绩','资格审核','面试','征文','培训','活动','直播','访谈']
POLICY_TERMS=DOC_TERMS+['认定办法','奖励办法','扶持办法','行动方案','目录']

OUTDIR=Path.home()/ 'Desktop' / '政策文件夹'
OUTDIR.mkdir(parents=True, exist_ok=True)
OUT_FORMAL=OUTDIR / '广东_2026_全部政策_正式政策.xlsx'
OUT_INTERP=OUTDIR / '广东_2026_全部政策_官方解读转载.xlsx'


def clean_html_text(s):
    return re.sub(r'\s+',' ',s or '').strip()

def strip_tags_keep_text(html):
    soup=BeautifulSoup(html,'html.parser')
    return clean_html_text(soup.get_text(' ',strip=True))

def resolve_sogou_url(href):
    if not href:
        return ''
    if href.startswith('http') and 'sogou.com/link' not in href:
        return href
    if href.startswith('/link?'):
        href='https://www.sogou.com'+href
    try:
        r=SESSION.get(href,timeout=20,allow_redirects=True)
        txt=r.text[:500]
        m=re.search(r'window\.location(?:\.replace)?\("([^"]+)"\)',txt)
        if m:
            return m.group(1)
        m=re.search(r'URL=([^"\']+)',txt)
        if m:
            return m.group(1)
        return r.url
    except Exception:
        return href


def sogou_search(query,page=1):
    url='https://www.sogou.com/web?query='+urllib.parse.quote(query)+f'&page={page}'
    r=SESSION.get(url,timeout=20)
    soup=BeautifulSoup(r.text,'html.parser')
    out=[]
    for box in soup.select('div.vrwrap, div.rb'):
        a=box.select_one('h3 a')
        if not a: continue
        title=strip_tags_keep_text(str(a))
        href=a.get('href','')
        summary=''
        date=''
        # summary containers vary
        for sel in ['.text-layout','.fz-mid','.str-text-info','.star-wiki','p','div']:
            el=box.select_one(sel)
            if el:
                t=clean_html_text(el.get_text(' ',strip=True))
                if len(t) > len(summary):
                    summary=t
        m=re.search(r'(20\d{2}[年\-/.]\d{1,2}[月\-/.]\d{1,2}日?)', summary)
        if m: date=m.group(1)
        out.append({'query':query,'search_page':page,'search_title':title,'search_summary':summary,'search_date':date,'search_url':href})
    return out


def normalize_date(text):
    if not text: return ''
    text=text.replace('发布时间：','').replace('发布日期：','').replace('时间：','')
    m=re.search(r'(20\d{2})[年\-/.](\d{1,2})[月\-/.](\d{1,2})', text)
    if not m: return ''
    y,mn,d=m.groups()
    return f'{y}-{int(mn):02d}-{int(d):02d}'


def fetch_page(url):
    try:
        r=SESSION.get(url,timeout=25)
        html=r.text
    except Exception:
        return {}
    soup=BeautifulSoup(html,'html.parser')
    title=''
    if soup.title: title=clean_html_text(soup.title.get_text(' ',strip=True))
    for sel in ['h1','.title','.article-title','.arti_title','.bt','meta[property="og:title"]']:
        el=soup.select_one(sel)
        if el:
            cand=el.get('content') if el.name=='meta' else el.get_text(' ',strip=True)
            cand=clean_html_text(cand)
            if len(cand) > len(title)/2:
                title=cand
                break
    text=clean_html_text(soup.get_text(' ',strip=True))
    date=''
    for pat in [r'发布时间[:：]?\s*(20\d{2}[年\-/.]\d{1,2}[月\-/.]\d{1,2}日?)',r'发布日期[:：]?\s*(20\d{2}[年\-/.]\d{1,2}[月\-/.]\d{1,2}日?)',r'(20\d{2}[年\-/.]\d{1,2}[月\-/.]\d{1,2}日?)']:
        m=re.search(pat,text[:6000])
        if m:
            date=normalize_date(m.group(1)); break
    summary=''
    desc=soup.select_one('meta[name="description"]')
    if desc and desc.get('content'): summary=clean_html_text(desc['content'])
    if not summary:
        summary=clean_html_text(text[:220])
    return {'page_title':title,'page_date':date,'page_text':text[:8000],'page_summary':summary,'final_url':r.url}


def domain(url):
    try:
        return urllib.parse.urlparse(url).netloc.lower()
    except Exception:
        return ''


def is_guangdong_related(row):
    url=row['url']; ttl=row['标题']; txt=row.get('正文','')
    dom=domain(url)
    if dom.endswith('.gd.gov.cn') or dom==PROV_DOMAIN: return True
    if any(dom.endswith(d) or dom==d for d in CITY_DOMAINS): return True
    gd_terms=['广东','广州','深圳','珠海','汕头','佛山','韶关','河源','梅州','惠州','汕尾','东莞','中山','江门','阳江','湛江','茂名','肇庆','清远','潮州','揭阳','云浮']
    return any(t in (ttl+txt[:500]) for t in gd_terms)


def source_level(row):
    dom=domain(row['url'])
    ttl=row['标题']
    txt=row.get('正文','')
    if dom.endswith('.gov.cn') or dom in CITY_DOMAINS or dom==PROV_DOMAIN:
        return 'A'
    if any(x in ttl+txt[:200] for x in ['政策解读','解读']) and (dom.endswith('.gov.cn') or 'gd.gov.cn' in dom):
        return 'B'
    if any(h in dom for h in ['people.com.cn','xinhuanet.com','southcn.com','ycwb.com']):
        return 'B'
    return 'C'


def classify_type(row):
    ttl=row['标题']
    txt=row.get('正文','')
    dom=domain(row['url'])
    combined=ttl+' '+txt[:500]
    if '解读' in ttl or '政策解读' in combined:
        return '官方解读'
    if not (dom.endswith('.gov.cn') or any(h in dom for h in ['people.com.cn','xinhuanet.com','southcn.com','ycwb.com'])):
        return '新闻报道'
    if any(term in combined for term in POLICY_TERMS):
        return '原始政策'
    return '新闻报道'


def policy_strength(row):
    ttl=row['标题']
    combined=ttl+' '+row.get('正文','')[:300]
    for t in POLICY_TERMS:
        if t in combined:
            return t
    return ''


def credibility_score(row):
    lvl=row['来源等级']
    s=40 if lvl=='A' else 28 if lvl=='B' else 0
    if is_guangdong_related(row): s+=20
    if row['发布时间'].startswith('2026-'): s+=20
    if row['文种']: s+=10
    if row['结果类型']=='原始政策': s+=10
    return min(s,100)


def dedup_key(row):
    title=re.sub(r'\s+','',row['标题'])
    return title[:80]

queries=[]
# layer1 global discovery
for t in ['通知','公告','实施意见','工作方案','若干措施','管理办法','实施细则','指南']:
    queries.append(('L1',f'2026年 广东 {t}'))
# layer2 official targeted supplement
for d in ['gd.gov.cn','gz.gov.cn','sz.gov.cn','foshan.gov.cn','dg.gov.cn','zhuhai.gov.cn','jiangmen.gov.cn','huizhou.gov.cn']:
    for t in ['通知','公告','实施意见','若干措施']:
        queries.append(('L2',f'2026年 广东 {t} site:{d}'))
# layer3 theme/department vertical
for theme in ['科技','财政','产业','人才','制造业','招商','农业','消费','能源','新能源汽车']:
    queries.append(('L3',f'2026年 广东 {theme} 通知'))
    queries.append(('L3',f'2026年 广东 {theme} 若干措施'))
for dept in ['发改委','财政厅','科技厅','工信厅','商务厅','农业农村厅','人社厅','教育厅']:
    queries.append(('L3',f'2026年 广东 {dept} 通知'))
queries.append(('L3','2026年 广东 政策解读'))
queries.append(('L3','2026年 广东 官方解读'))

raw=[]
for layer,q in queries:
    try:
        for p in [1]:
            items=sogou_search(q,page=p)
            for it in items:
                it['layer']=layer
                raw.append(it)
        time.sleep(0.6)
    except Exception as e:
        print('search_error',q,e,file=sys.stderr)

print('raw results',len(raw))

# resolve and fetch
seen_urls=OrderedDict()
for item in raw:
    resolved=resolve_sogou_url(item['search_url'])
    if not resolved.startswith('http'): continue
    if resolved not in seen_urls:
        seen_urls[resolved]=item
    else:
        seen_urls[resolved]['query'] += ' | ' + item['query']

print('unique resolved',len(seen_urls))

rows=[]
for i,(url,it) in enumerate(seen_urls.items(),1):
    page=fetch_page(url)
    title=page.get('page_title') or it['search_title']
    date=page.get('page_date') or normalize_date(it.get('search_date',''))
    text=page.get('page_text','')
    summary=page.get('page_summary') or it['search_summary']
    row={
        '标题':title,
        '正文':text,
        '摘要':summary,
        '来源地方':'广东',
        '来源网站':domain(url),
        '原文地址':page.get('final_url') or url,
        'url':page.get('final_url') or url,
        '发布时间':date,
        '地域':'广东',
        '年份':'2026',
        '主题':'全部政策',
        'query':it['query'],
        'layer':it['layer'],
    }
    rows.append(row)
    if i % 20 == 0:
        print('fetched',i)
    time.sleep(0.4)

print('fetched rows',len(rows))

# filter/classify
filtered=[]
for row in rows:
    row['标题']=clean_html_text(row['标题'])
    if not row['标题'] or len(row['标题'])<6: continue
    if not is_guangdong_related(row):
        continue
    if not row['发布时间'] or not row['发布时间'].startswith('2026-'):
        continue
    if any(b in row['标题'] for b in BAD_TITLE_HINTS):
        continue
    row['来源等级']=source_level(row)
    if row['来源等级']=='C':
        continue
    row['文种']=policy_strength(row)
    row['结果类型']=classify_type(row)
    if row['结果类型']=='新闻报道' and row['来源等级']!='B':
        continue
    if row['结果类型']=='原始政策' and not row['文种']:
        continue
    row['可信度评分']=credibility_score(row)
    row['备注']=f"{row['layer']}召回；来源等级{row['来源等级']}"
    filtered.append(row)

print('after filter',len(filtered))

# dedupe by URL then title
best_by_key={}
for row in filtered:
    key=row['url']
    if key not in best_by_key or row['可信度评分']>best_by_key[key]['可信度评分']:
        best_by_key[key]=row
filtered=list(best_by_key.values())

best_by_title={}
for row in filtered:
    key=dedup_key(row)
    if key not in best_by_title or row['可信度评分']>best_by_title[key]['可信度评分']:
        best_by_title[key]=row
filtered=list(best_by_title.values())
print('after dedup',len(filtered))

formal=[]; interp=[]
for row in filtered:
    if row['结果类型']=='原始政策' and row['来源等级']=='A':
        formal.append(row)
    elif row['结果类型'] in ['官方解读','新闻报道'] or row['来源等级']=='B':
        interp.append(row)

formal=sorted(formal,key=lambda r:(-r['可信度评分'],r['发布时间'],r['标题']))
interp=sorted(interp,key=lambda r:(-r['可信度评分'],r['发布时间'],r['标题']))

cols=['标题','摘要','来源地方','来源网站','原文地址','发布时间','结果类型','地域','年份','主题','文种','可信度评分','备注','正文']

def write_xlsx(path, rows, cols):
    wb=Workbook()
    ws=wb.active
    ws.title='Sheet1'
    ws.append(cols)
    for row in rows:
        ws.append([row.get(c,'') for c in cols])
    for idx,col in enumerate(cols,1):
        max_len=max([len(str(col))]+[len(str(r.get(col,''))) for r in rows[:200]]) if rows else len(col)
        ws.column_dimensions[get_column_letter(idx)].width=min(max(max_len+2,12),50)
    wb.save(path)

write_xlsx(OUT_FORMAL, formal, cols)
write_xlsx(OUT_INTERP, interp, cols)

summary={
    'raw_count':len(raw),
    'unique_urls':len(seen_urls),
    'filtered_count':len(filtered),
    'formal_count':len(formal),
    'interp_count':len(interp),
    'formal_path':str(OUT_FORMAL),
    'interp_path':str(OUT_INTERP),
}
print(json.dumps(summary,ensure_ascii=False,indent=2))

# save debug json
Path('/tmp/policy_guangdong_2026_summary.json').write_text(json.dumps(summary,ensure_ascii=False,indent=2),encoding='utf-8')
Path('/tmp/policy_guangdong_2026_formal.json').write_text(json.dumps(formal,ensure_ascii=False,indent=2),encoding='utf-8')
Path('/tmp/policy_guangdong_2026_interp.json').write_text(json.dumps(interp,ensure_ascii=False,indent=2),encoding='utf-8')
